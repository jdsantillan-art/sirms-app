"""
Export Views for SIRMS
Handles Excel export functionality for reports
"""
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from .models import IncidentReport, DOSchedule, CounselingSchedule, CaseEvaluation


@login_required
def export_all_reports_excel(request):
    """
    Export all reports to Excel for DO and authorized users
    """
    # Check permissions
    if request.user.role not in ['do', 'counselor', 'principal']:
        messages.error(request, 'You do not have permission to export reports.')
        return redirect('dashboard')
    
    # Get all reports
    reports = IncidentReport.objects.all().select_related(
        'incident_type', 'reported_student', 'reporter', 'classification'
    ).order_by('-created_at')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "All Reports"
    
    # Define styles
    header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Case ID', 'Status', 'Student Name', 'Student Gender', 'Grade', 'Section',
        'Incident Type', 'Type Category', 'Incident Date', 'Incident Time',
        'Reporter Name', 'Reporter Role', 'Description', 'Classification',
        'Reported Date', 'Days Open'
    ]
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write data
    for row_idx, report in enumerate(reports, start=2):
        # Calculate days open
        days_open = (datetime.now().date() - report.created_at.date()).days
        
        # Get student name
        student_name = report.reported_student.get_full_name() if report.reported_student else report.involved_students
        
        # Get classification
        classification = ""
        if hasattr(report, 'classification'):
            classification = report.classification.get_severity_display()
        
        # Get type category
        type_category = ""
        if report.incident_type:
            type_category = "Prohibited Acts" if report.incident_type.severity == 'prohibited' else "School Policies"
        
        data = [
            report.case_id,
            report.get_status_display(),
            student_name,
            report.student_gender.title() if report.student_gender else 'Not specified',
            f"Grade {report.grade_level}",
            report.section_name,
            report.incident_type.name if report.incident_type else 'Not specified',
            type_category,
            report.incident_date.strftime('%Y-%m-%d') if report.incident_date else '',
            report.incident_time.strftime('%H:%M') if report.incident_time else '',
            f"{report.reporter_first_name} {report.reporter_last_name}",
            report.reporter.get_role_display() if report.reporter else '',
            report.description[:100] + '...' if len(report.description) > 100 else report.description,
            classification,
            report.created_at.strftime('%Y-%m-%d %H:%M'),
            days_open
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Adjust column widths
    column_widths = {
        'A': 12,  # Case ID
        'B': 15,  # Status
        'C': 20,  # Student Name
        'D': 12,  # Gender
        'E': 10,  # Grade
        'F': 20,  # Section
        'G': 30,  # Incident Type
        'H': 18,  # Type Category
        'I': 12,  # Incident Date
        'J': 10,  # Incident Time
        'K': 20,  # Reporter Name
        'L': 15,  # Reporter Role
        'M': 40,  # Description
        'N': 15,  # Classification
        'O': 18,  # Reported Date
        'P': 10,  # Days Open
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'SIRMS_All_Reports_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response



@login_required
def export_behavior_concerns_excel(request):
    """
    Export completed behavior concerns (DO handled cases) to Excel
    """
    # Check permissions - only DO can export
    if request.user.role != 'do':
        messages.error(request, 'You do not have permission to export behavior concerns.')
        return redirect('dashboard')
    
    # Get completed behavior concerns (resolved status)
    reports = IncidentReport.objects.filter(
        status='resolved'
    ).select_related(
        'incident_type', 'reported_student', 'reporter', 'classification'
    ).order_by('-created_at')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Completed Behavior Concerns"
    
    # Define styles
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")  # Green
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Case ID', 'Student Name', 'Student Gender', 'Grade', 'Section',
        'Incident Type', 'Type Category', 'Incident Date', 'Incident Time',
        'Reporter Name', 'Reporter Role', 'Description', 
        'Classification', 'Reported Date', 'Completed Date', 'Days to Complete',
        'Scheduled Appointments', 'Appointment Details', 'Final Notes'
    ]
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write data
    for row_idx, report in enumerate(reports, start=2):
        # Calculate days to complete
        days_to_complete = (report.updated_at.date() - report.created_at.date()).days
        
        # Get student name
        student_name = report.reported_student.get_full_name() if report.reported_student else report.involved_students
        
        # Get classification
        classification = ""
        if hasattr(report, 'classification') and report.classification:
            classification = report.classification.get_severity_display()
        
        # Get type category
        type_category = ""
        if report.incident_type:
            type_category = "Prohibited Acts" if report.incident_type.severity == 'prohibited' else "School Policies"
        
        # Get scheduled appointments
        schedules = DOSchedule.objects.filter(report=report).order_by('scheduled_date')
        appointment_count = schedules.count()
        appointment_details = []
        
        for schedule in schedules:
            detail = f"{schedule.get_schedule_type_display()} on {schedule.scheduled_date.strftime('%Y-%m-%d %H:%M')}"
            if schedule.location:
                detail += f" at {schedule.location}"
            if schedule.status:
                detail += f" ({schedule.get_status_display()})"
            appointment_details.append(detail)
        
        appointment_details_str = "\n".join(appointment_details) if appointment_details else "No appointments scheduled"
        
        # Get final notes from latest schedule
        final_notes = ""
        if schedules.exists():
            latest_schedule = schedules.last()
            if latest_schedule.notes:
                final_notes = latest_schedule.notes
        
        data = [
            report.case_id,
            student_name,
            report.student_gender.title() if report.student_gender else 'Not specified',
            f"Grade {report.grade_level}",
            report.section_name,
            report.incident_type.name if report.incident_type else 'Not specified',
            type_category,
            report.incident_date.strftime('%Y-%m-%d') if report.incident_date else '',
            report.incident_time.strftime('%H:%M') if report.incident_time else '',
            f"{report.reporter_first_name} {report.reporter_last_name}",
            report.reporter.get_role_display() if report.reporter else '',
            report.description,
            classification,
            report.created_at.strftime('%Y-%m-%d %H:%M'),
            report.updated_at.strftime('%Y-%m-%d %H:%M'),
            days_to_complete,
            appointment_count,
            appointment_details_str,
            final_notes
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Adjust column widths
    column_widths = {
        'A': 12,  # Case ID
        'B': 20,  # Student Name
        'C': 12,  # Gender
        'D': 10,  # Grade
        'E': 20,  # Section
        'F': 30,  # Incident Type
        'G': 18,  # Type Category
        'H': 12,  # Incident Date
        'I': 10,  # Incident Time
        'J': 20,  # Reporter Name
        'K': 15,  # Reporter Role
        'L': 40,  # Description
        'M': 15,  # Classification
        'N': 18,  # Reported Date
        'O': 18,  # Completed Date
        'P': 12,  # Days to Complete
        'Q': 12,  # Scheduled Appointments
        'R': 40,  # Appointment Details
        'S': 40,  # Final Notes
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add summary at the bottom
    last_row = len(reports) + 2
    ws.cell(row=last_row + 1, column=1, value="SUMMARY").font = Font(bold=True, size=12)
    ws.cell(row=last_row + 2, column=1, value=f"Total Completed Cases: {reports.count()}")
    ws.cell(row=last_row + 3, column=1, value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=last_row + 4, column=1, value=f"Exported By: {request.user.get_full_name()}")
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'SIRMS_Completed_Behavior_Concerns_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response



@login_required
def export_completed_reports_excel(request):
    """
    Export completed counseling reports to Excel for Guidance Counselors
    """
    # Check permissions - only counselors and guidance can export
    if request.user.role not in ['counselor', 'guidance']:
        messages.error(request, 'You do not have permission to export completed reports.')
        return redirect('dashboard')
    
    # Get all completed counseling sessions
    completed_sessions = CounselingSchedule.objects.filter(
        counselor=request.user,
        status='completed'
    ).select_related(
        'student', 'evaluation', 'evaluation__report', 'evaluation__report__incident_type', 
        'evaluation__report__reporter', 'counselor'
    ).order_by('-completed_date')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Completed Counseling Reports"
    
    # Define styles
    header_fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")  # Green
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Case ID', 'Student Name', 'Student Gender', 'Grade', 'Section',
        'Incident Type', 'Type Category', 'Incident Date', 
        'Reporter Name', 'Reporter Role',
        'Session Date', 'Session Time', 'Completed Date', 'Completed Time',
        'Days to Complete', 'Location', 'Counselor', 
        'Session Notes', 'Recommendations', 'Follow-up Required'
    ]
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write data
    for row_idx, session in enumerate(completed_sessions, start=2):
        if not session.evaluation or not session.evaluation.report:
            continue
            
        report = session.evaluation.report
        evaluation = session.evaluation
        
        # Calculate days to complete
        if session.completed_date and report.created_at:
            days_to_complete = (session.completed_date.date() - report.created_at.date()).days
        else:
            days_to_complete = 0
        
        # Get student name
        student_name = session.student.get_full_name() if session.student else 'N/A'
        
        # Get type category
        type_category = ""
        if report.incident_type:
            type_category = "Prohibited Acts" if report.incident_type.severity == 'prohibited' else "School Policies"
        
        # Get recommendations
        recommendations = evaluation.recommendations if hasattr(evaluation, 'recommendations') else ''
        
        # Follow-up required
        follow_up = 'Yes' if evaluation.requires_follow_up else 'No' if hasattr(evaluation, 'requires_follow_up') else 'N/A'
        
        data = [
            report.case_id,
            student_name,
            report.student_gender.title() if report.student_gender else 'Not specified',
            f"Grade {report.grade_level}",
            report.section_name,
            report.incident_type.name if report.incident_type else 'Not specified',
            type_category,
            report.incident_date.strftime('%Y-%m-%d') if report.incident_date else '',
            f"{report.reporter_first_name} {report.reporter_last_name}",
            report.reporter.get_role_display() if report.reporter else '',
            session.scheduled_date.strftime('%Y-%m-%d') if session.scheduled_date else '',
            session.scheduled_date.strftime('%H:%M') if session.scheduled_date else '',
            session.completed_date.strftime('%Y-%m-%d') if session.completed_date else '',
            session.completed_date.strftime('%H:%M') if session.completed_date else '',
            days_to_complete,
            session.location or 'Not specified',
            session.counselor.get_full_name(),
            session.notes or 'No notes',
            recommendations,
            follow_up
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Adjust column widths
    column_widths = {
        'A': 12,  # Case ID
        'B': 20,  # Student Name
        'C': 12,  # Gender
        'D': 10,  # Grade
        'E': 20,  # Section
        'F': 30,  # Incident Type
        'G': 18,  # Type Category
        'H': 12,  # Incident Date
        'I': 20,  # Reporter Name
        'J': 15,  # Reporter Role
        'K': 12,  # Session Date
        'L': 10,  # Session Time
        'M': 12,  # Completed Date
        'N': 10,  # Completed Time
        'O': 12,  # Days to Complete
        'P': 20,  # Location
        'Q': 20,  # Counselor
        'R': 40,  # Session Notes
        'S': 40,  # Recommendations
        'T': 15,  # Follow-up Required
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add summary at the bottom
    last_row = len(list(completed_sessions)) + 2
    ws.cell(row=last_row + 1, column=1, value="SUMMARY").font = Font(bold=True, size=12)
    ws.cell(row=last_row + 2, column=1, value=f"Total Completed Sessions: {completed_sessions.count()}")
    ws.cell(row=last_row + 3, column=1, value=f"Counselor: {request.user.get_full_name()}")
    ws.cell(row=last_row + 4, column=1, value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws.cell(row=last_row + 5, column=1, value=f"Exported By: {request.user.get_full_name()}")
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'SIRMS_Completed_Counseling_Reports_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response
