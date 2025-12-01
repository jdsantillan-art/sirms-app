from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.utils import timezone
from django.http import JsonResponse
from django.db import models
from datetime import datetime, timedelta
import json
from .models import (
    CustomUser, IncidentReport, Classification, CounselingSession,
    Notification, IncidentType, Section, TeacherAssignment,
    Curriculum, Track, Grade, ViolationHistory, CaseEvaluation,
    InternalNote, SystemBackup, ReportAnalytics, LegalReference, CounselingSchedule
)
from .forms import (
    CustomUserCreationForm, IncidentReportForm, ClassificationForm,
    CounselingSessionForm, CaseEvaluationForm,
    InternalNoteForm, CurriculumForm, TrackForm, GradeForm,
    SectionForm, TeacherAssignmentForm, IncidentTypeForm, ReportAnalyticsForm
)


def notify_adviser_of_counseling(report, scheduled_date, location, counselor_name, schedule_type='Counseling'):
    """
    Notify the adviser/teacher of a student when counseling is scheduled.
    Finds the teacher based on curriculum, grade, and section from the report.
    """
    if not report:
        return
    
    # Try to find the teacher/adviser based on the report's academic details
    teachers_to_notify = []
    
    # Method 1: Find by teacher_name field in the report
    if report.teacher_name:
        # Try to find teacher by name
        teacher_users = CustomUser.objects.filter(
            role='teacher',
            first_name__icontains=report.teacher_name.split()[0] if report.teacher_name.split() else report.teacher_name
        )
        if teacher_users.exists():
            teachers_to_notify.extend(teacher_users)
    
    # Method 2: Find by Section adviser (if section is linked)
    if report.section and report.section.adviser:
        if report.section.adviser not in teachers_to_notify:
            teachers_to_notify.append(report.section.adviser)
    
    # Method 3: Find by TeacherAssignment matching curriculum, grade, section
    if report.curriculum and report.grade_level and report.section_name:
        teacher_assignments = TeacherAssignment.objects.filter(
            curriculum=report.curriculum,
            grade_level=report.grade_level,
            section_name__iexact=report.section_name
        )
        
        for assignment in teacher_assignments:
            # Try to find the teacher user by name
            teacher_name_parts = assignment.teacher_name.split()
            if teacher_name_parts:
                matching_teachers = CustomUser.objects.filter(
                    role='teacher',
                    first_name__icontains=teacher_name_parts[0]
                )
                if len(teacher_name_parts) > 1:
                    matching_teachers = matching_teachers.filter(
                        last_name__icontains=teacher_name_parts[-1]
                    )
                
                for teacher in matching_teachers:
                    if teacher not in teachers_to_notify:
                        teachers_to_notify.append(teacher)
    
    # Send notifications to all found teachers
    student_name = report.reported_student.get_full_name() if report.reported_student else report.involved_students
    
    for teacher in teachers_to_notify:
        Notification.objects.create(
            user=teacher,
            title=f'{schedule_type} Session Scheduled for Your Student',
            message=f'A {schedule_type.lower()} session has been scheduled for your student {student_name} (Case: {report.case_id}) on {scheduled_date.strftime("%B %d, %Y at %I:%M %p")}. Location: {location if location else "TBA"}. Counselor: {counselor_name}.',
            report=report
        )

def home(request):
    return render(request, 'home.html')

def health_check(request):
    """Simple health check endpoint for Render"""
    from django.http import JsonResponse
    return JsonResponse({'status': 'ok', 'service': 'sirms'})

def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            
            # Auto-create Counselor entry for ESP teachers
            if user.role == 'esp_teacher':
                from .models import Counselor
                # Check if counselor entry already exists
                if not Counselor.objects.filter(name=user.get_full_name()).exists():
                    Counselor.objects.create(
                        name=user.get_full_name(),
                        email=user.email,
                        specialization='ESP Teacher/VPF Coordinator',
                        is_active=True
                    )
            
            username = user.username
            raw_password = form.cleaned_data.get('password1')
            auth_user = authenticate(request, username=username, password=raw_password)
            if auth_user is not None:
                login(request, auth_user)
                next_url = request.POST.get('next') or request.GET.get('next')
                if next_url:
                    return redirect(next_url)
                return redirect('dashboard')
            else:
                messages.error(request, 'Registration succeeded, but auto-login failed. Please log in.')
                return redirect('login')
        else:
            # Handle specific error messages
            if 'email' in form.errors:
                messages.error(request, 'This email address is already registered.')
            if 'password2' in form.errors:
                for error in form.errors['password2']:
                    if 'match' in error.lower():
                        messages.error(request, 'Passwords do not match.')
                    else:
                        messages.error(request, error)
            if 'password1' in form.errors:
                for error in form.errors['password1']:
                    messages.error(request, error)
            # Handle other field errors
            for field, errors in form.errors.items():
                if field not in ['email', 'password1', 'password2']:
                    for error in errors:
                        messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = CustomUserCreationForm()
    return render(request, 'register.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            next_url = request.POST.get('next') or request.GET.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password')
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('home')

@login_required
def dashboard(request):
    user = request.user
    context = {'user_role': user.role}
    
    # Get current date and determine school year (June to May)
    current_date = timezone.now()
    current_month = current_date.month
    
    # School year starts in June
    if current_month >= 6:  # June onwards - current school year
        school_year_start = current_date.replace(month=6, day=1)
    else:  # Before June - previous school year
        school_year_start = current_date.replace(year=current_date.year - 1, month=6, day=1)
    
    # Generate trend data for school year months (June to May)
    school_months = ['June', 'July', 'Aug', 'Sept', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar', 'Apr', 'May']
    trend_data = []
    
    for i, month_name in enumerate(school_months):
        # Calculate the actual month number
        month_num = (6 + i) if (6 + i) <= 12 else (6 + i - 12)
        year_offset = 0 if (6 + i) <= 12 else 1
        
        month_start = school_year_start.replace(month=month_num, day=1)
        if year_offset == 1:
            month_start = month_start.replace(year=month_start.year + 1)
        
        # Get last day of month
        if month_num == 12:
            month_end = month_start.replace(day=31)
        else:
            next_month = month_start.replace(month=month_num + 1) if month_num < 12 else month_start.replace(year=month_start.year + 1, month=1)
            month_end = next_month - timedelta(days=1)
        
        month_reports = IncidentReport.objects.filter(
            created_at__range=[month_start, month_end]
        ).count()
        
        trend_data.append({
            'month': month_name,
            'reports': month_reports
        })
    
    # Grade data - for Reports by Grade Level
    grade_data = []
    for grade in range(7, 13):
        count = IncidentReport.objects.filter(grade_level=str(grade)).count()
        grade_data.append({
            'grade': f'Grade {grade}',
            'count': count
        })
    
    # Violation type data - Prohibited Acts vs Other School Policies (for PIE CHART)
    prohibited_count = IncidentReport.objects.filter(
        incident_type__severity='prohibited'
    ).count()
    
    school_policy_count = IncidentReport.objects.filter(
        incident_type__severity='school_policy'
    ).count()
    
    violation_type_data = [
        {'name': 'Prohibited Acts', 'value': prohibited_count},
        {'name': 'Other School Policies', 'value': school_policy_count}
    ]
    
    # Stacked data for compatibility
    stacked_data = []
    for i, month_name in enumerate(school_months):
        month_num = (6 + i) if (6 + i) <= 12 else (6 + i - 12)
        year_offset = 0 if (6 + i) <= 12 else 1
        
        month_start = school_year_start.replace(month=month_num, day=1)
        if year_offset == 1:
            month_start = month_start.replace(year=month_start.year + 1)
        
        if month_num == 12:
            month_end = month_start.replace(day=31)
        else:
            next_month = month_start.replace(month=month_num + 1) if month_num < 12 else month_start.replace(year=month_start.year + 1, month=1)
            month_end = next_month - timedelta(days=1)
        
        prohibited_count_month = IncidentReport.objects.filter(
            created_at__range=[month_start, month_end],
            incident_type__severity='prohibited'
        ).count()
        
        school_policy_count_month = IncidentReport.objects.filter(
            created_at__range=[month_start, month_end],
            incident_type__severity='school_policy'
        ).count()
        
        stacked_data.append({
            'month': month_name,
            'major': prohibited_count_month,
            'minor': school_policy_count_month
        })
    
    # Add analytics data to context
    context.update({
        'trend_data': json.dumps(trend_data),
        'grade_data': json.dumps(grade_data),
        'violation_type_data': json.dumps(violation_type_data),
        'stacked_data': json.dumps(stacked_data),
        'resolution_data': json.dumps([]),  # Default empty for all roles
    })
    
    if user.role == 'student':
        reports = IncidentReport.objects.filter(reported_student=user)
        context.update({
            'pending_reports': reports.filter(status='pending').count(),
            'under_review': reports.filter(status='under_review').count(),
            'resolved': reports.filter(status__in=['resolved', 'closed']).count(),
            'recent_reports': reports.order_by('-created_at')[:5],
            'counseling_sessions': CounselingSession.objects.filter(student=user, status='scheduled'),
        })
    elif user.role == 'teacher':
        reports = IncidentReport.objects.filter(reporter=user)
        context.update({
            'total_reports': reports.count(),
            'pending': reports.filter(status='pending').count(),
            'recent_reports': reports.order_by('-created_at')[:5],
        })
    elif user.role == 'do':
        reports = IncidentReport.objects.all()
        
        # Count reports with classifications
        minor_cases = reports.filter(classification__severity='minor').count()
        major_cases = reports.filter(classification__severity='major').count()
        
        # If no classifications exist, count by status instead
        if minor_cases == 0 and major_cases == 0:
            # Count reports that would typically be minor (handled by DO)
            minor_cases = reports.filter(
                status__in=['pending', 'under_review', 'resolved']
            ).exclude(
                status='referred_to_counselor'
            ).count()
            
            # Count reports referred to counselor (major cases)
            major_cases = reports.filter(
                status='referred_to_counselor'
            ).count()
        
        context.update({
            'total_reports': reports.count(),
            'pending': reports.filter(status='pending').count(),
            'classified': reports.filter(status='classified').count(),
            'minor_cases_count': minor_cases,
            'major_cases_count': major_cases,
            'recent_reports': reports.order_by('-created_at')[:10],
        })
    elif user.role == 'counselor':
        major_cases = IncidentReport.objects.filter(classification__severity='major')
        upcoming_sessions = CounselingSession.objects.filter(
            counselor=user,
            status='scheduled',
            scheduled_date__gte=timezone.now()
        ).order_by('scheduled_date')[:5]
        
        # Resolution data for counselor
        resolution_data = []
        for i in range(12):
            month_start = start_date + timedelta(days=30*i)
            month_end = month_start + timedelta(days=30)
            
            resolved_count = IncidentReport.objects.filter(
                created_at__range=[month_start, month_end],
                status='resolved'
            ).count()
            
            pending_count = IncidentReport.objects.filter(
                created_at__range=[month_start, month_end],
                status__in=['pending', 'under_review', 'classified', 'evaluated']
            ).count()
            
            resolution_data.append({
                'month': month_start.strftime('%b'),
                'resolved': resolved_count,
                'pending': pending_count
            })
        
        # Calculate real counseling success rate
        total_sessions = CounselingSession.objects.filter(counselor=user).count()
        completed_sessions = CounselingSession.objects.filter(
            counselor=user, 
            status='completed'
        ).count()
        success_rate = int((completed_sessions / total_sessions * 100)) if total_sessions > 0 else 0
        
        # New analytics metrics - Fixed counts
        from .models import VPFCase
        
        # Count PA and OSP based on incident type severity
        total_prohibited_acts = IncidentReport.objects.filter(
            incident_type__severity='prohibited'
        ).count()
        
        total_osp = IncidentReport.objects.filter(
            incident_type__severity='school_policy'
        ).count()
        
        # Scheduled sessions - from CounselingSchedule, not CounselingSession
        scheduled_sessions_count = CounselingSchedule.objects.filter(
            counselor=user,
            status='scheduled'
        ).count()
        
        # Completed sessions - from CounselingSchedule with status='completed'
        completed_sessions_count = CounselingSchedule.objects.filter(
            counselor=user,
            status='completed'
        ).count()
        
        # Completed VPF - based on VPFCase status='completed'
        completed_vpf = VPFCase.objects.filter(status='completed').count()
        
        # Total VPF Referrals - VPF cases assigned by this counselor
        total_vpf_referrals = VPFCase.objects.filter(assigned_by=user).count()
        
        context.update({
            'total_prohibited_acts': total_prohibited_acts,
            'total_osp': total_osp,
            'scheduled_sessions': scheduled_sessions_count,
            'completed_vpf': completed_vpf,
            'total_vpf_referrals': total_vpf_referrals,
            'counseling_success_rate': success_rate,
            'completed_sessions': completed_sessions_count,
            'major_cases': major_cases.count(),
            'pending_evaluations': major_cases.filter(status='classified').count(),
            'recent_cases': major_cases.order_by('-created_at')[:5],
            'upcoming_sessions': upcoming_sessions,
            'resolution_data': json.dumps(resolution_data),
        })
    elif user.role == 'principal':
        reports = IncidentReport.objects.all()
        total_reports = reports.count()
        resolved_reports = reports.filter(status='resolved').count()
        resolution_rate = int((resolved_reports / total_reports * 100)) if total_reports > 0 else 0
        
        # Resolution data for principal
        resolution_data = []
        for i in range(12):
            month_start = start_date + timedelta(days=30*i)
            month_end = month_start + timedelta(days=30)
            
            resolved_count = IncidentReport.objects.filter(
                created_at__range=[month_start, month_end],
                status='resolved'
            ).count()
            
            pending_count = IncidentReport.objects.filter(
                created_at__range=[month_start, month_end],
                status__in=['pending', 'under_review', 'classified', 'evaluated']
            ).count()
            
            resolution_data.append({
                'month': month_start.strftime('%b'),
                'resolved': resolved_count,
                'pending': pending_count
            })
        
        context.update({
            'total_reports': total_reports,
            'resolution_rate': resolution_rate,
            'active_sanctions': Sanction.objects.filter(
                issued_at__gte=timezone.now() - timedelta(days=30),
                sanction_type__in=['suspension', 'expulsion']
            ).count(),
            'scheduled_sessions': CounselingSession.objects.filter(status='scheduled').count(),
            'repeat_offenders': CaseEvaluation.objects.filter(is_repeat_offender=True).count(),
            'total_students': CustomUser.objects.filter(role='student').count(),
            'active_teachers': CustomUser.objects.filter(role='teacher', is_active=True).count(),
            'active_cases': reports.exclude(status__in=['resolved', 'closed']).count(),
            'recent_reports': reports.order_by('-created_at')[:5],
            'resolution_data': json.dumps(resolution_data),
        })
    elif user.role == 'esp_teacher':
        # ESP Teacher Dashboard Analytics
        from .models import VPFCase, VPFSchedule, Counselor
        
        # Find the Counselor record that matches this ESP teacher's name
        esp_teacher_name = user.get_full_name()
        matching_counselors = Counselor.objects.filter(name__icontains=esp_teacher_name)
        
        if matching_counselors.exists():
            # Get VPF cases assigned to this ESP teacher
            vpf_cases_assigned = VPFCase.objects.filter(
                esp_teacher_assigned__in=matching_counselors
            )
            
            # Get scheduled sessions for this ESP teacher
            scheduled_vpf_sessions = VPFSchedule.objects.filter(
                vpf_case__esp_teacher_assigned__in=matching_counselors,
                status='scheduled'
            ).count()
            
            # Get completed VPF cases
            completed_vpf_count = vpf_cases_assigned.filter(status='completed').count()
            pending_vpf_count = vpf_cases_assigned.filter(status__in=['pending', 'ongoing']).count()
            
            # Total VPF referrals assigned to this ESP teacher
            total_vpf_referrals_count = vpf_cases_assigned.count()
            
            # VPF Improvement Rate Data (monthly trend)
            vpf_improvement_data = []
            for i in range(12):
                month_start = start_date + timedelta(days=30*i)
                month_end = month_start + timedelta(days=30)
                
                # Count completed VPF in this month
                completed_in_month = vpf_cases_assigned.filter(
                    updated_at__range=[month_start, month_end],
                    status='completed'
                ).count()
                
                # Count pending/ongoing VPF in this month
                pending_in_month = vpf_cases_assigned.filter(
                    assigned_at__range=[month_start, month_end],
                    status__in=['pending', 'ongoing']
                ).count()
                
                # Calculate improvement rate
                total_in_month = completed_in_month + pending_in_month
                improvement_rate = int((completed_in_month / total_in_month * 100)) if total_in_month > 0 else 0
                
                vpf_improvement_data.append({
                    'month': month_start.strftime('%b'),
                    'completed': completed_in_month,
                    'pending': pending_in_month,
                    'rate': improvement_rate
                })
            
            context.update({
                'scheduled_vpf_sessions': scheduled_vpf_sessions,
                'completed_vpf': completed_vpf_count,
                'total_vpf_referrals': total_vpf_referrals_count,
                'pending_vpf': pending_vpf_count,
                'ongoing_vpf': vpf_cases_assigned.filter(status='ongoing').count(),
                'vpf_improvement_data': json.dumps(vpf_improvement_data),
            })
        else:
            # No matching counselor record found
            context.update({
                'scheduled_vpf_sessions': 0,
                'completed_vpf': 0,
                'total_vpf_referrals': 0,
                'pending_vpf': 0,
                'ongoing_vpf': 0,
                'vpf_improvement_data': json.dumps([]),
            })
    elif user.role in ['guidance', 'maintenance']:
        context.update({
            'total_users': CustomUser.objects.count(),
            'active_teachers': CustomUser.objects.filter(role='teacher', is_active=True).count(),
            'incident_types_count': IncidentType.objects.count(),
            'legal_references_count': LegalReference.objects.count(),
        })
    
    notifications = Notification.objects.filter(user=user, is_read=False).order_by('-created_at')[:5]
    context['notifications'] = notifications
    context['unread_count'] = Notification.objects.filter(user=user, is_read=False).count()
    
    return render(request, 'dashboard.html', context)

@login_required
def analytics_dashboard(request):
    """Dedicated analytics dashboard with comprehensive charts and insights"""
    user = request.user
    context = {'user_role': user.role}
    
    # Get current date and determine school year (June to May)
    current_date = timezone.now()
    current_month = current_date.month
    
    # School year starts in June
    if current_month >= 6:  # June onwards - current school year
        school_year_start = current_date.replace(month=6, day=1)
    else:  # Before June - previous school year
        school_year_start = current_date.replace(year=current_date.year - 1, month=6, day=1)
    
    # Generate trend data for school year months (June to May)
    school_months = ['June', 'July', 'Aug', 'Sept', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar', 'Apr', 'May']
    trend_data = []
    
    for i, month_name in enumerate(school_months):
        # Calculate the actual month number
        month_num = (6 + i) if (6 + i) <= 12 else (6 + i - 12)
        year_offset = 0 if (6 + i) <= 12 else 1
        
        month_start = school_year_start.replace(month=month_num, day=1)
        if year_offset == 1:
            month_start = month_start.replace(year=month_start.year + 1)
        
        # Get last day of month
        if month_num == 12:
            month_end = month_start.replace(day=31)
        else:
            next_month = month_start.replace(month=month_num + 1) if month_num < 12 else month_start.replace(year=month_start.year + 1, month=1)
            month_end = next_month - timedelta(days=1)
        
        month_reports = IncidentReport.objects.filter(
            created_at__range=[month_start, month_end]
        ).count()
        
        trend_data.append({
            'month': month_name,
            'reports': month_reports
        })
    
    # Grade data - for Reports by Grade Level
    grade_data = []
    for grade in range(7, 13):
        count = IncidentReport.objects.filter(grade_level=str(grade)).count()
        grade_data.append({
            'grade': f'Grade {grade}',
            'count': count
        })
    
    # Violation type data - Prohibited Acts vs Other School Policies (for PIE CHART)
    prohibited_count = IncidentReport.objects.filter(
        incident_type__severity='prohibited'
    ).count()
    
    school_policy_count = IncidentReport.objects.filter(
        incident_type__severity='school_policy'
    ).count()
    
    violation_type_data = [
        {'name': 'Prohibited Acts', 'value': prohibited_count},
        {'name': 'Other School Policies', 'value': school_policy_count}
    ]
    
    # Stacked data for Major vs Minor (kept for compatibility)
    stacked_data = []
    for i, month_name in enumerate(school_months):
        month_num = (6 + i) if (6 + i) <= 12 else (6 + i - 12)
        year_offset = 0 if (6 + i) <= 12 else 1
        
        month_start = school_year_start.replace(month=month_num, day=1)
        if year_offset == 1:
            month_start = month_start.replace(year=month_start.year + 1)
        
        if month_num == 12:
            month_end = month_start.replace(day=31)
        else:
            next_month = month_start.replace(month=month_num + 1) if month_num < 12 else month_start.replace(year=month_start.year + 1, month=1)
            month_end = next_month - timedelta(days=1)
        
        prohibited_count_month = IncidentReport.objects.filter(
            created_at__range=[month_start, month_end],
            incident_type__severity='prohibited'
        ).count()
        
        school_policy_count_month = IncidentReport.objects.filter(
            created_at__range=[month_start, month_end],
            incident_type__severity='school_policy'
        ).count()
        
        stacked_data.append({
            'month': month_name,
            'major': prohibited_count_month,
            'minor': school_policy_count_month
        })
    
    # Add analytics data to context
    context.update({
        'trend_data': json.dumps(trend_data),
        'grade_data': json.dumps(grade_data),
        'violation_type_data': json.dumps(violation_type_data),
        'stacked_data': json.dumps(stacked_data),
        'resolution_data': json.dumps([]),  # Default empty for all roles
    })
    
    # Role-specific data
    if user.role == 'do':
        reports = IncidentReport.objects.all()
        minor_cases = reports.filter(classification__severity='minor').count()
        major_cases = reports.filter(classification__severity='major').count()
        
        if minor_cases == 0 and major_cases == 0:
            minor_cases = reports.filter(
                status__in=['pending', 'under_review', 'resolved']
            ).exclude(status='referred_to_counselor').count()
            major_cases = reports.filter(status='referred_to_counselor').count()
        
        # Classified today
        today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        classified_today = reports.filter(
            classification__classified_at__gte=today_start
        ).count()
        
        context.update({
            'total_reports': reports.count(),
            'pending': reports.filter(status='pending').count(),
            'classified_today': classified_today,
            'minor_cases_count': minor_cases,
            'major_cases_count': major_cases,
            'recent_reports': reports.order_by('-created_at')[:10],
        })
    
    elif user.role == 'counselor':
        major_cases = IncidentReport.objects.filter(classification__severity='major')
        upcoming_sessions = CounselingSession.objects.filter(
            counselor=user,
            status='scheduled',
            scheduled_date__gte=timezone.now()
        ).order_by('scheduled_date')[:5]
        
        # Resolution data for counselor
        resolution_data = []
        for i in range(12):
            month_start = start_date + timedelta(days=30*i)
            month_end = month_start + timedelta(days=30)
            
            resolved_count = IncidentReport.objects.filter(
                created_at__range=[month_start, month_end],
                status='resolved'
            ).count()
            
            pending_count = IncidentReport.objects.filter(
                created_at__range=[month_start, month_end],
                status__in=['pending', 'under_review', 'classified', 'evaluated']
            ).count()
            
            resolution_data.append({
                'month': month_start.strftime('%b'),
                'resolved': resolved_count,
                'pending': pending_count
            })
        
        # Calculate counseling success rate
        total_sessions = CounselingSession.objects.filter(counselor=user).count()
        completed_sessions = CounselingSession.objects.filter(
            counselor=user, 
            status='completed'
        ).count()
        success_rate = int((completed_sessions / total_sessions * 100)) if total_sessions > 0 else 0
        
        # Scheduled sessions from CounselingSchedule
        scheduled_sessions_count = CounselingSchedule.objects.filter(
            counselor=user,
            status='scheduled'
        ).count()
        
        context.update({
            'major_cases': major_cases.count(),
            'scheduled_sessions': scheduled_sessions_count,
            'counseling_success_rate': success_rate,
            'pending_evaluations': major_cases.filter(status='classified').count(),
            'recent_cases': major_cases.order_by('-created_at')[:5],
            'upcoming_sessions': upcoming_sessions,
            'resolution_data': json.dumps(resolution_data),
        })
    
    elif user.role == 'principal':
        reports = IncidentReport.objects.all()
        total_reports = reports.count()
        resolved_reports = reports.filter(status='resolved').count()
        resolution_rate = int((resolved_reports / total_reports * 100)) if total_reports > 0 else 0
        
        # Resolution data for principal
        resolution_data = []
        for i in range(12):
            month_start = start_date + timedelta(days=30*i)
            month_end = month_start + timedelta(days=30)
            
            resolved_count = IncidentReport.objects.filter(
                created_at__range=[month_start, month_end],
                status='resolved'
            ).count()
            
            pending_count = IncidentReport.objects.filter(
                created_at__range=[month_start, month_end],
                status__in=['pending', 'under_review', 'classified', 'evaluated']
            ).count()
            
            resolution_data.append({
                'month': month_start.strftime('%b'),
                'resolved': resolved_count,
                'pending': pending_count
            })
        
        context.update({
            'total_reports': total_reports,
            'resolution_rate': resolution_rate,
            'active_sanctions': Sanction.objects.filter(
                issued_at__gte=timezone.now() - timedelta(days=30),
                sanction_type__in=['suspension', 'expulsion']
            ).count(),
            'repeat_offenders': CaseEvaluation.objects.filter(is_repeat_offender=True).count(),
            'total_students': CustomUser.objects.filter(role='student').count(),
            'active_teachers': CustomUser.objects.filter(role='teacher', is_active=True).count(),
            'active_cases': reports.exclude(status__in=['resolved', 'closed']).count(),
            'resolution_data': json.dumps(resolution_data),
        })
    
    return render(request, 'analytics_dashboard.html', context)

@login_required
def report_incident(request):
    if request.method == 'POST':
        form = IncidentReportForm(request.POST, request.FILES)
        if form.is_valid():
            # Create the report manually since we're using custom fields
            report = IncidentReport()
            report.reporter = request.user
            
            # Set reporter information
            report.reporter_first_name = form.cleaned_data['reporter_first_name']
            report.reporter_middle_name = form.cleaned_data['reporter_middle_name']
            report.reporter_last_name = form.cleaned_data['reporter_last_name']
            
            # Set academic information
            report.curriculum = form.cleaned_data['curriculum']
            report.grade_level = form.cleaned_data['grade_level']
            report.section_name = form.cleaned_data['section_name']
            report.teacher_name = form.cleaned_data['teacher_name']
            
            # Set other fields
            report.involved_students = form.cleaned_data['involved_students']
            report.student_gender = form.cleaned_data.get('student_gender', '')
            report.incident_date = form.cleaned_data['incident_date']
            report.incident_time = form.cleaned_data['incident_time']
            report.incident_type = form.cleaned_data['incident_type']
            
            # Handle bullying type if provided
            bullying_type = request.POST.get('bullying_type', '')
            if bullying_type:
                report.bullying_type = bullying_type
            
            report.description = form.cleaned_data['description']
            
            report.evidence = form.cleaned_data['evidence']
            
            # Try to find and assign the reported student from involved_students
            involved_students_text = form.cleaned_data['involved_students']
            if involved_students_text:
                # Try to find a student by email, username, or name in the involved students text
                # Split by common delimiters
                import re
                potential_identifiers = re.split(r'[,;\n]+', involved_students_text)
                
                for identifier in potential_identifiers:
                    identifier = identifier.strip()
                    if identifier:
                        # Try to find student by email
                        student = CustomUser.objects.filter(
                            role='student',
                            email__iexact=identifier
                        ).first()
                        
                        # If not found by email, try username
                        if not student:
                            student = CustomUser.objects.filter(
                                role='student',
                                username__iexact=identifier
                            ).first()
                        
                        # If not found by username, try to match by full name
                        if not student:
                            # Split the identifier into potential first and last names
                            name_parts = identifier.split()
                            if len(name_parts) >= 2:
                                first_name = name_parts[0]
                                last_name = name_parts[-1]
                                
                                # Try exact match on first and last name
                                student = CustomUser.objects.filter(
                                    role='student',
                                    first_name__iexact=first_name,
                                    last_name__iexact=last_name
                                ).first()
                                
                                # If still not found, try partial match
                                if not student:
                                    student = CustomUser.objects.filter(
                                        role='student',
                                        first_name__icontains=first_name,
                                        last_name__icontains=last_name
                                    ).first()
                        
                        # If found, assign as reported_student
                        if student:
                            report.reported_student = student
                            break
            
            report.save()
            
            # Create notification for all Discipline Officers
            do_users = CustomUser.objects.filter(role='do')
            for do_user in do_users:
                Notification.objects.create(
                    user=do_user,
                    title='New Incident Report Submitted',
                    message=f'New incident report {report.case_id} filed by {request.user.get_full_name()}. Requires fact-checking and classification.',
                    report=report
                )
            
            # Notify the reporter
            Notification.objects.create(
                user=request.user,
                title='Report Submitted Successfully',
                message=f'Your incident report {report.case_id} has been submitted and is being reviewed by the Discipline Office.',
                report=report
            )
            
            messages.success(request, f'Report {report.case_id} submitted successfully')
            return redirect('my_reports')
        else:
            # Display form errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")
    else:
        form = IncidentReportForm()
        # Pre-fill reporter information
        form.fields['reporter_first_name'].initial = request.user.first_name
        form.fields['reporter_middle_name'].initial = request.user.middle_name if hasattr(request.user, 'middle_name') else ''
        form.fields['reporter_last_name'].initial = request.user.last_name
    
    # Get context data for the template
    context = {
        'form': form,
        'incident_types': IncidentType.objects.all().order_by('severity', 'name'),
        'teacher_assignments': TeacherAssignment.objects.all(),
    }
    
    return render(request, 'report_incident.html', context)

@login_required
def my_reports(request):
    # Show reports created by the current user (regardless of role)
    reports = IncidentReport.objects.filter(reporter=request.user).order_by('-created_at')
    
    return render(request, 'my_reports.html', {
        'reports': reports
    })

@login_required
def all_reports(request):
    if request.user.role not in ['do', 'counselor', 'principal']:
        return redirect('dashboard')
    
    reports = IncidentReport.objects.all().select_related('classification', 'incident_type', 'reported_student').order_by('-incident_date', '-incident_time')
    
    # Calculate statistics
    pending_count = reports.filter(status='pending').count()
    classified_count = reports.filter(status='classified').count()
    resolved_count = reports.filter(status__in=['resolved', 'closed']).count()
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        reports = reports.filter(status=status_filter)
    
    # Filter by severity (minor/major)
    severity_filter = request.GET.get('severity')
    if severity_filter:
        reports = reports.filter(classification__severity=severity_filter)
    
    return render(request, 'all_reports.html', {
        'reports': reports,
        'pending_count': pending_count,
        'classified_count': classified_count,
        'resolved_count': resolved_count,
    })

@login_required
def report_detail(request, case_id):
    report = get_object_or_404(IncidentReport, case_id=case_id)
    
    # Check permissions - allow users to view reports they created or are involved in
    if request.user.role == 'student':
        # Students can view reports where they are the reported student or the reporter
        if report.reported_student != request.user and report.reporter != request.user:
            return redirect('dashboard')
    elif request.user.role not in ['do', 'counselor', 'principal']:
        # Other roles can only view reports they created
        if report.reporter != request.user:
            return redirect('dashboard')
    
    context = {'report': report}
    
    # Add classification if exists
    if hasattr(report, 'classification'):
        context['classification'] = report.classification
    
    # Add counseling sessions related to this report
    counseling_sessions = CounselingSession.objects.filter(
        student=report.reported_student,
        report=report
    ).order_by('scheduled_date')
    context['counseling_sessions'] = counseling_sessions
    
    # Add internal notes for authorized users
    if request.user.role in ['do', 'counselor', 'principal']:
        internal_notes = InternalNote.objects.filter(
            report=report
        ).order_by('-created_at')
        context['internal_notes'] = internal_notes
    
    if request.user.role == 'do' and request.method == 'POST' and 'classify' in request.POST:
        form = ClassificationForm(request.POST)
        if form.is_valid():
            classification = form.save(commit=False)
            classification.report = report
            classification.classified_by = request.user
            classification.save()
            report.status = 'classified'
            report.save()
            
            # Notify student
            Notification.objects.create(
                user=report.reported_student,
                title='Case Classified',
                message=f'Your case {report.case_id} has been classified as {classification.get_severity_display()}',
                report=report
            )
            
            # If major case, notify all counselors
            if classification.severity == 'major':
                counselor_users = CustomUser.objects.filter(role='counselor')
                for counselor in counselor_users:
                    Notification.objects.create(
                        user=counselor,
                        title='New Major Case for Evaluation',
                        message=f'Major case {report.case_id} requires counseling evaluation. Student: {report.reported_student.get_full_name()}',
                        report=report
                    )
            
            messages.success(request, 'Report classified successfully')
            return redirect('report_detail', case_id=case_id)
    
    if request.user.role == 'do':
        context['classification_form'] = ClassificationForm()
    
    return render(request, 'report_detail.html', context)

@login_required
def counseling_schedule(request):
    if request.user.role == 'student':
        sessions = CounselingSession.objects.filter(student=request.user).order_by('-scheduled_date')
    elif request.user.role == 'teacher':
        # Show sessions for students reported by this teacher
        sessions = CounselingSession.objects.filter(
            report__reporter=request.user
        ).order_by('-scheduled_date')
    elif request.user.role == 'counselor':
        sessions = CounselingSession.objects.filter(counselor=request.user).order_by('-scheduled_date')
    else:
        sessions = CounselingSession.objects.all().order_by('-scheduled_date')
    
    return render(request, 'counseling_schedule.html', {'sessions': sessions})

@login_required
def create_sanction(request, case_id):
    if request.user.role != 'principal':
        return redirect('dashboard')
    
    report = get_object_or_404(IncidentReport, case_id=case_id)
    
    if request.method == 'POST':
        form = SanctionForm(request.POST)
        if form.is_valid():
            sanction = form.save(commit=False)
            sanction.report = report
            sanction.issued_by = request.user
            sanction.save()
            report.status = 'sanctioned'
            report.save()
            
            # Notify student and counselor
            Notification.objects.create(
                user=report.reported_student,
                title='Sanction Issued',
                message=f'A {sanction.get_sanction_type_display()} has been issued for case {report.case_id}',
                report=report
            )
            
            messages.success(request, 'Sanction issued successfully')
            return redirect('report_detail', case_id=case_id)
    else:
        form = SanctionForm()
    
    return render(request, 'create_sanction.html', {'form': form, 'report': report})

@login_required
def notifications(request):
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    
    if request.method == 'POST':
        notification_id = request.POST.get('notification_id')
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.is_read = True
        notification.save()
    
    return render(request, 'notifications.html', {'notifications': notifications})

@login_required
def schedule_notification_detail(request, notification_id):
    """View for schedule notification details only (not full report)"""
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    
    # Mark as read
    if not notification.is_read:
        notification.is_read = True
        notification.save()
    
    context = {
        'notification': notification,
        'is_schedule_notification': True
    }
    
    return render(request, 'schedule_notification_detail.html', context)

@login_required
def account_settings(request):
    if request.method == 'POST':
        request.user.email = request.POST.get('email')
        request.user.first_name = request.POST.get('first_name')
        request.user.last_name = request.POST.get('last_name')
        request.user.save()
        messages.success(request, 'Account settings updated')
        return redirect('account_settings')
    
    return render(request, 'account_settings.html')

# API endpoints for dynamic dropdowns
@login_required
def get_tracks(request):
    curriculum_id = request.GET.get('curriculum_id')
    tracks = Track.objects.filter(curriculum_id=curriculum_id).values('id', 'name')
    return JsonResponse(list(tracks), safe=False)

@login_required
def get_grades(request):
    track_id = request.GET.get('track_id')
    grades = Grade.objects.filter(track_id=track_id).values('id', 'level')
    return JsonResponse(list(grades), safe=False)

@login_required
def get_sections(request):
    grade_id = request.GET.get('grade_id')
    sections = Section.objects.filter(grade_id=grade_id).values('id', 'name')
    return JsonResponse(list(sections), safe=False)

@login_required
def get_teachers(request):
    grade_level = request.GET.get('grade_level')
    track_code = request.GET.get('track_code')
    section_name = request.GET.get('section_name')
    
    teachers = TeacherAssignment.objects.filter(
        grade_level=grade_level,
        track_code=track_code,
        section_name=section_name
    ).values('id', 'teacher_name')
    
    return JsonResponse(list(teachers), safe=False)

# Student-specific views
@login_required
def violation_history(request):
    if request.user.role != 'student':
        return redirect('dashboard')
    
    history = ViolationHistory.objects.filter(student=request.user)
    return render(request, 'student/violation_history.html', {'history': history})

@login_required
def legal_references(request):
    incident_types = IncidentType.objects.all()
    return render(request, 'shared/legal_references.html', {'incident_types': incident_types})

# Teacher-specific views
@login_required
def advisee_records(request):
    if request.user.role != 'teacher':
        return redirect('dashboard')
    
    # Get sections advised by this teacher
    advised_sections = request.user.advised_sections.all()
    
    # Get all students in the teacher's advised sections
    advisee_students = CustomUser.objects.filter(
        role='student',
        section__in=[section.name for section in advised_sections]
    )
    
    # Get all incident reports involving the teacher's advisees
    advisee_reports = IncidentReport.objects.filter(
        reported_student__in=advisee_students
    ).select_related(
        'reported_student', 'incident_type', 'reporter'
    ).prefetch_related(
        'counseling_sessions', 'classification'
    ).order_by('-created_at')
    
    # Also get reports where the teacher's section is mentioned in section_name field
    section_names = [section.name for section in advised_sections]
    additional_reports = IncidentReport.objects.filter(
        section_name__in=section_names
    ).exclude(
        id__in=advisee_reports.values_list('id', flat=True)
    ).select_related(
        'incident_type', 'reporter'
    ).order_by('-created_at')
    
    # Combine and sort all reports
    all_reports = list(advisee_reports) + list(additional_reports)
    all_reports.sort(key=lambda x: x.created_at, reverse=True)
    
    # Get counseling sessions referred back to teacher
    referred_sessions = CounselingSession.objects.filter(
        student__in=advisee_students,
        status='referred_to_teacher'
    ).select_related('student', 'report').order_by('-scheduled_date')
    
    # Calculate statistics
    total_incidents = len(all_reports)
    pending_incidents = len([r for r in all_reports if r.status == 'pending'])
    resolved_incidents = len([r for r in all_reports if r.status == 'resolved'])
    
    # Update section violation counts
    section_stats = {}
    for section in advised_sections:
        section_reports = [r for r in all_reports if (
            (hasattr(r, 'reported_student') and r.reported_student and r.reported_student.section == section.name) or
            r.section_name == section.name
        )]
        section_stats[section.name] = {
            'total': len(section_reports),
            'pending': len([r for r in section_reports if r.status == 'pending']),
            'resolved': len([r for r in section_reports if r.status == 'resolved'])
        }
    
    return render(request, 'teacher/advisee_records.html', {
        'advised_sections': advised_sections,
        'advisee_reports': all_reports,
        'referred_sessions': referred_sessions,
        'section_stats': section_stats,
        'total_incidents': total_incidents,
        'pending_incidents': pending_incidents,
        'resolved_incidents': resolved_incidents,
    })

# DO-specific views
@login_required
def fact_check_reports(request):
    if request.user.role != 'do':
        return redirect('dashboard')
    
    if request.method == 'POST':
        report_id = request.POST.get('report_id')
        action = request.POST.get('action')
        
        if report_id and action:
            report = get_object_or_404(IncidentReport, id=report_id)
            
            if action == 'verify':
                # Get classification details from POST data
                severity = request.POST.get('severity')
                notes = request.POST.get('notes', '')
                evidence_status = request.POST.get('evidence_status', 'clear')
                evidence_notes = request.POST.get('evidence_notes', '')
                student_id = request.POST.get('student_id')
                
                # Assign student if provided
                if student_id:
                    try:
                        student = CustomUser.objects.get(id=student_id, role='student')
                        report.reported_student = student
                    except CustomUser.DoesNotExist:
                        pass
                
                # Update evidence status
                report.evidence_status = evidence_status
                report.evidence_notes = evidence_notes
                
                # If evidence is insufficient, notify reporter and don't proceed
                if evidence_status == 'insufficient':
                    report.save()
                    
                    # Notify reporter about insufficient evidence
                    Notification.objects.create(
                        user=report.reporter,
                        title='Additional Evidence Required',
                        message=f'Case {report.case_id} requires more evidence. Please provide additional documentation or information. Reason: {evidence_notes}',
                        report=report
                    )
                    
                    messages.warning(request, f'Case {report.case_id} marked as needing more evidence. Reporter has been notified.')
                    return redirect('fact_check_reports')
                
                # If evidence is clear, proceed with classification
                if severity:
                    # Check if major case requires a student
                    if severity == 'major' and not report.reported_student:
                        messages.error(request, f'Cannot classify case {report.case_id} as Major: No student assigned. Please identify and add the student first (update Involved Students field with student email/username).')
                        return redirect('fact_check_reports')
                    
                    # Create classification
                    Classification.objects.create(
                        report=report,
                        classified_by=request.user,
                        severity=severity,
                        internal_notes=notes
                    )
                    
                    # Update report status
                    report.status = 'classified'
                    report.save()
                    
                    # Create violation history entry
                    if report.reported_student:
                        ViolationHistory.objects.create(
                            student=report.reported_student,
                            report=report,
                            violation_type=report.incident_type,
                            severity=severity,
                            date_occurred=report.incident_date
                        )
                    
                    # Send notifications based on routing decision
                    if severity == 'minor':
                        # Case stays with DO - notify DO staff
                        do_users = CustomUser.objects.filter(role='do')
                        for do_user in do_users:
                            if do_user != request.user:  # Don't notify the current user
                                Notification.objects.create(
                                    user=do_user,
                                    title='Case Assigned to DO',
                                    message=f'Case {report.case_id} ({report.incident_type.name}) has been assigned to Discipline Office for handling.',
                                    report=report
                                )
                        
                        # Notify student
                        if report.reported_student:
                            Notification.objects.create(
                                user=report.reported_student,
                                title='Case Under DO Review',
                                message=f'Your case {report.case_id} is being handled by the Discipline Office. You will be contacted for further action.',
                                report=report
                            )
                        
                        messages.success(request, f'Case {report.case_id} has been verified and assigned to Discipline Office.')
                    else:
                        # Send to guidance counselor
                        counselors = CustomUser.objects.filter(role='counselor')
                        for counselor in counselors:
                            Notification.objects.create(
                                user=counselor,
                                title='Case Referred for Counseling',
                                message=f'Case {report.case_id} ({report.incident_type.name}) has been referred to Guidance for counseling intervention.',
                                report=report
                            )
                        
                        # Notify student
                        if report.reported_student:
                            Notification.objects.create(
                                user=report.reported_student,
                                title='Counseling Session Required',
                                message=f'Your case {report.case_id} has been referred to the Guidance Counselor. A counseling session will be scheduled.',
                                report=report
                            )
                        
                        messages.success(request, f'Case {report.case_id} has been verified and sent to Guidance Counselor.')
                else:
                    # Just verify without classification (old behavior)
                    report.status = 'under_review'
                    report.save()
                    
                    counselors = CustomUser.objects.filter(role='counselor')
                    for counselor in counselors:
                        Notification.objects.create(
                            user=counselor,
                            title='Report Verified - Ready for Review',
                            message=f'Case {report.case_id} has been fact-checked and is ready for review',
                            report=report
                        )
                    
                    messages.success(request, f'Case {report.case_id} has been verified and moved to review.')
                
            elif action == 'request_info':
                # Create notification for reporter
                Notification.objects.create(
                    user=report.reporter,
                    title='Additional Information Requested',
                    message=f'More information is needed for Case {report.case_id}. Please provide additional details.',
                    report=report
                )
                
                messages.info(request, f'Additional information requested for Case {report.case_id}.')
        
        return redirect('fact_check_reports')
    
    # GET request - display reports
    pending_reports = IncidentReport.objects.filter(
        status='pending'
    ).select_related(
        'reporter', 'incident_type', 'reported_student'
    ).order_by('-created_at')
    
    # Calculate statistics
    total_pending = pending_reports.count()
    today_reports = pending_reports.filter(created_at__date=timezone.now().date()).count()
    urgent_reports = pending_reports.filter(incident_type__severity='major').count()
    
    # Get all students for the dropdown
    students = CustomUser.objects.filter(role='student', is_active=True).order_by('first_name', 'last_name')
    
    return render(request, 'do/fact_check_reports.html', {
        'reports': pending_reports,
        'students': students,
        'total_pending': total_pending,
        'today_reports': today_reports,
        'urgent_reports': urgent_reports,
    })

@login_required
def behavior_concerns(request):
    """Handle behavior concerns - cases routed to DO for handling"""
    if request.user.role != 'do':
        return redirect('dashboard')
    
    if request.method == 'POST':
        report_id = request.POST.get('report_id')
        action = request.POST.get('action')
        
        if report_id and action:
            report = get_object_or_404(IncidentReport, id=report_id)
            
            if action == 'schedule_appointment':
                appointment_type = request.POST.get('appointment_type')
                scheduled_date_str = request.POST.get('scheduled_date')
                location = request.POST.get('location', '')
                notes = request.POST.get('notes', '')
                student_id = request.POST.get('student_id')
                
                if not appointment_type or not scheduled_date_str:
                    messages.error(request, 'Please provide appointment type and date.')
                    return redirect('behavior_concerns')
                
                # Parse and make timezone-aware
                from datetime import datetime
                scheduled_date = datetime.fromisoformat(scheduled_date_str)
                if scheduled_date.tzinfo is None:
                    scheduled_date = timezone.make_aware(scheduled_date)
                
                # Get student object
                student = None
                if student_id:
                    student = get_object_or_404(CustomUser, id=student_id)
                elif report.reported_student:
                    student = report.reported_student
                
                # Create DOSchedule entry for sidebar display
                do_schedule = DOSchedule.objects.create(
                    report=report,
                    discipline_officer=request.user,
                    student=student,
                    schedule_type=appointment_type,
                    scheduled_date=scheduled_date,
                    location=location if location else 'Discipline Office',
                    purpose=f"Behavior concern follow-up for case {report.case_id}",
                    notes=notes,
                    status='scheduled'
                )
                
                # Create internal note for the appointment
                appointment_note = f"DO Appointment Scheduled\nType: {appointment_type}\nDate: {scheduled_date.strftime('%B %d, %Y at %I:%M %p')}\nLocation: {location if location else 'DO Office'}\nNotes: {notes if notes else 'N/A'}"
                
                InternalNote.objects.create(
                    report=report,
                    author=request.user,
                    note=appointment_note,
                    is_private=False
                )
                
                # Add to classification notes
                if hasattr(report, 'classification'):
                    current_notes = report.classification.internal_notes or ''
                    timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
                    updated_notes = f"{current_notes}\n\n[{timestamp}] {appointment_note}".strip()
                    report.classification.internal_notes = updated_notes
                    report.classification.save()
                
                # Notify student
                if student:
                    Notification.objects.create(
                        user=student,
                        title=f'DO Appointment Scheduled - Case {report.case_id}',
                        message=f'You have been scheduled for a {appointment_type} with the Discipline Office.\n\nDate & Time: {scheduled_date.strftime("%B %d, %Y at %I:%M %p")}\nLocation: {location if location else "DO Office"}\n\nPlease be on time. If you cannot attend, contact the DO office immediately.',
                        report=report
                    )
                
                messages.success(request, f'Appointment scheduled for {scheduled_date.strftime("%B %d, %Y at %I:%M %p")}. Student has been notified and added to DO Schedule.')
                return redirect('behavior_concerns')
            
            elif action == 'update_status':
                new_status = request.POST.get('status')
                evaluation_action = request.POST.get('evaluation_action', '')
                notes = request.POST.get('notes', '')
                
                if not notes or not evaluation_action:
                    messages.error(request, 'Please provide evaluation action and notes.')
                    return redirect('behavior_concerns')
                
                # Update report status
                old_status = report.status
                report.status = new_status
                
                # Add evaluation notes to classification with timestamp
                if hasattr(report, 'classification'):
                    current_notes = report.classification.internal_notes or ''
                    timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
                    status_display = report.get_status_display()
                    updated_notes = f"{current_notes}\n\n[{timestamp}] DO Evaluation by {request.user.get_full_name()}\nAction: {evaluation_action}\nStatus: {status_display}\nNotes: {notes}".strip()
                    report.classification.internal_notes = updated_notes
                    report.classification.save()
                
                report.save()
                
                # Create internal note for tracking
                InternalNote.objects.create(
                    report=report,
                    author=request.user,
                    note=f"DO Evaluation - {evaluation_action}: {notes}",
                    is_private=False
                )
                
                # Create detailed notification message
                action_messages = {
                    'Intake Interview': 'The Discipline Office has scheduled an intake interview regarding your case. Please report to the DO office as instructed.',
                    'Investigate': 'The Discipline Office is investigating your case. You may be called for questioning or clarification.',
                    'Parent Conference': 'A parent conference has been scheduled regarding your case. Your parent/guardian will be contacted by the Discipline Office.'
                }
                action_msg = action_messages.get(evaluation_action, 'The Discipline Office is taking action on your case.')
                
                # Notify reporter about evaluation
                Notification.objects.create(
                    user=report.reporter,
                    title=f'DO Evaluation - Case {report.case_id}',
                    message=f'The Discipline Office has evaluated case {report.case_id}.\n\nAction Taken: {evaluation_action}\nStatus: {report.get_status_display()}\n\nNotes: {notes}',
                    report=report
                )
                
                # Notify student (violator) about evaluation
                if report.reported_student:
                    Notification.objects.create(
                        user=report.reported_student,
                        title=f'Behavioral Concern Evaluation - Case {report.case_id}',
                        message=f'The Discipline Office has evaluated your case {report.case_id}.\n\nAction: {evaluation_action}\n\n{action_msg}\n\nPlease check with the Discipline Office for further instructions.',
                        report=report
                    )
                
                messages.success(request, f'Case {report.case_id} evaluated with action "{evaluation_action}". Student and reporter have been notified.')
        
        return redirect('behavior_concerns')
    
    # Get only cases routed to DO (severity='minor' means handled by DO)
    do_cases = IncidentReport.objects.filter(
        classification__severity='minor'  # 'minor' = Handle by DO
    ).select_related(
        'classification', 'incident_type', 'reporter', 'reported_student'
    ).order_by('-created_at')
    
    # Calculate statistics
    total_cases = do_cases.count()
    pending_cases = do_cases.filter(status='classified').count()
    completed_cases = do_cases.filter(status='resolved').count()
    
    return render(request, 'do/behavior_concerns.html', {
        'reports': do_cases,
        'total_classified': total_cases,
        'minor_cases': total_cases,
        'major_cases': 0,
        'total_cases': total_cases,
        'pending_cases': pending_cases,
        'completed_cases': completed_cases,
    })

# Keep old function name for backward compatibility
classify_violations = behavior_concerns

@login_required
def pre_counseling_schedule(request):
    if request.user.role != 'do':
        return redirect('dashboard')
    
    if request.method == 'POST':
        case_id = request.POST.get('case_id')
        counselor_name = request.POST.get('counselor_name')
        scheduled_date = request.POST.get('scheduled_date')
        notes = request.POST.get('notes', '')
        
        try:
            report = get_object_or_404(IncidentReport, id=case_id)
            
            # Validate that report has a student
            if not report.reported_student:
                messages.error(request, 'This report does not have a student assigned. Please assign a student first.')
                return redirect('pre_counseling_schedule')
            
            counselor = None
            
            # Get assigned counselor user if they exist in CustomUser
            # This maintains backward compatibility with existing counselor users
            if counselor_name:
                # Try to find a user with role counselor matching the name
                counselor_users = CustomUser.objects.filter(role='counselor', is_active=True)
                for cu in counselor_users:
                    if cu.get_full_name() == counselor_name:
                        counselor = cu
                        break
            
            # Parse the scheduled date
            from datetime import datetime, timedelta
            scheduled_datetime = datetime.strptime(scheduled_date, '%Y-%m-%dT%H:%M')
            
            # CHECK FOR SCHEDULE CONFLICTS
            conflicts = []
            
            # Check counselor conflicts (within 1 hour window)
            counselor_sessions = CounselingSession.objects.filter(
                remarks__contains=f"Counselor: {counselor_name}",
                status='scheduled'
            )
            
            for session in counselor_sessions:
                time_diff = abs((session.scheduled_date - scheduled_datetime).total_seconds())
                if time_diff < 3600:  # Within 1 hour
                    conflicts.append(f"⚠️ Counselor {counselor_name} has another session at {session.scheduled_date.strftime('%I:%M %p')}")
            
            # Check student conflicts (within 1 hour window)
            student_sessions = CounselingSession.objects.filter(
                student=report.reported_student,
                status='scheduled'
            )
            
            for session in student_sessions:
                time_diff = abs((session.scheduled_date - scheduled_datetime).total_seconds())
                if time_diff < 3600:  # Within 1 hour
                    conflicts.append(f"⚠️ Student {report.reported_student.get_full_name()} has another session at {session.scheduled_date.strftime('%I:%M %p')}")
            
            # If conflicts found, show warning but allow scheduling
            if conflicts:
                for conflict in conflicts:
                    messages.warning(request, conflict)
                messages.warning(request, 'Schedule conflict detected! Please verify the time or reschedule.')
            
            # Create counseling session with counselor name in remarks
            session_remarks = f"Counselor: {counselor_name}"
            if notes:
                session_remarks += f"\nNotes: {notes}"
            
            session = CounselingSession.objects.create(
                report=report,
                student=report.reported_student,
                counselor=counselor,
                scheduled_date=scheduled_datetime,
                status='scheduled',
                remarks=session_remarks
            )
            
            # Notify the student
            Notification.objects.create(
                user=report.reported_student,
                title='Pre-Counseling Session Scheduled',
                message=f'A pre-counseling session has been scheduled with {counselor_name} on {session.scheduled_date.strftime("%B %d, %Y at %I:%M %p")} regarding case {report.case_id}.',
                report=report
            )
            
            # Notify the reporter (if teacher)
            if report.reporter and report.reporter.role == 'teacher':
                Notification.objects.create(
                    user=report.reporter,
                    title='Pre-Counseling Session Scheduled',
                    message=f'A pre-counseling session has been scheduled for student {report.reported_student.get_full_name()} with {counselor_name} on {session.scheduled_date.strftime("%B %d, %Y at %I:%M %p")} regarding case {report.case_id}.',
                    report=report
                )
            
            # Notify the assigned counselor if they have a user account
            if counselor:
                Notification.objects.create(
                    user=counselor,
                    title='New Pre-Counseling Session Assigned',
                    message=f'You have been assigned a pre-counseling session with {report.reported_student.get_full_name()} on {session.scheduled_date.strftime("%B %d, %Y at %I:%M %p")} for case {report.case_id}.',
                    report=report
                )
            
            # Update report status
            report.status = 'under_review'
            report.save()
            
            messages.success(request, f'Pre-counseling session scheduled successfully for {report.reported_student.get_full_name()} with {counselor_name}')
            return redirect('pre_counseling_schedule')
            
        except ValueError as e:
            messages.error(request, f'Invalid date format. Please use the date picker.')
            return redirect('pre_counseling_schedule')
        except Exception as e:
            messages.error(request, f'Error scheduling session: {str(e)}')
            return redirect('pre_counseling_schedule')
    
    # Get minor cases that need pre-counseling (only those with a student assigned)
    minor_cases = IncidentReport.objects.filter(
        classification__severity='minor',
        status='classified',
        reported_student__isnull=False  # Only show cases with a student
    ).select_related('reported_student', 'incident_type', 'classification')
    
    # Get all counselors for assignment from the Counselor model
    from .models import Counselor
    counselors = Counselor.objects.filter(is_active=True).order_by('name')
    
    # Get scheduled sessions
    scheduled_sessions = CounselingSession.objects.filter(
        status='scheduled'
    ).select_related('student', 'counselor', 'report').order_by('scheduled_date')
    
    return render(request, 'do/pre_counseling_schedule.html', {
        'cases': minor_cases,
        'counselors': counselors,
        'scheduled_sessions': scheduled_sessions,
    })

@login_required
def case_history(request):
    if request.user.role not in ['do', 'counselor', 'principal']:
        return redirect('dashboard')
    
    # If counselor, show VPF cases instead
    if request.user.role == 'counselor':
        from .models import VPFCase
        vpf_cases = VPFCase.objects.select_related(
            'student', 'report', 'assigned_by'
        ).order_by('-assigned_at')
        
        # Filter by status
        status_filter = request.GET.get('status', '')
        if status_filter:
            vpf_cases = vpf_cases.filter(status=status_filter)
        
        return render(request, 'counselor/for_vpf.html', {
            'vpf_cases': vpf_cases,
            'total_cases': vpf_cases.count(),
            'pending_cases': vpf_cases.filter(status='pending').count(),
            'scheduled_cases': vpf_cases.filter(status='scheduled').count(),
            'completed_cases': vpf_cases.filter(status='completed').count(),
        })
    
    # Handle Excel export
    if request.GET.get('export') == 'excel':
        return export_reports_to_excel(request)
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    severity_filter = request.GET.get('severity', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')
    
    # Get all incident reports with related data
    reports = IncidentReport.objects.select_related(
        'reporter', 'incident_type', 'reported_student', 'classification', 'sanction'
    ).prefetch_related(
        'counseling_sessions'
    ).order_by('-created_at')
    
    # Apply filters
    if status_filter:
        reports = reports.filter(status=status_filter)
    
    if severity_filter and severity_filter != 'all':
        reports = reports.filter(classification__severity=severity_filter)
    
    if date_from:
        try:
            from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
            reports = reports.filter(incident_date__gte=from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
            reports = reports.filter(incident_date__lte=to_date)
        except ValueError:
            pass
    
    if search_query:
        reports = reports.filter(
            Q(case_id__icontains=search_query) |
            Q(reported_student__first_name__icontains=search_query) |
            Q(reported_student__last_name__icontains=search_query) |
            Q(reporter__first_name__icontains=search_query) |
            Q(reporter__last_name__icontains=search_query) |
            Q(incident_type__name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Calculate statistics
    total_reports = reports.count()
    pending_reports = reports.filter(status='pending').count()
    resolved_reports = reports.filter(status='resolved').count()
    this_month_reports = reports.filter(
        created_at__month=timezone.now().month,
        created_at__year=timezone.now().year
    ).count()
    
    # Get unique values for filters
    status_choices = IncidentReport.STATUS_CHOICES
    incident_types = IncidentType.objects.all()
    
    return render(request, 'shared/case_history.html', {
        'reports': reports,
        'total_reports': total_reports,
        'pending_reports': pending_reports,
        'resolved_reports': resolved_reports,
        'this_month_reports': this_month_reports,
        'status_choices': status_choices,
        'incident_types': incident_types,
        'current_filters': {
            'status': status_filter,
            'severity': severity_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search_query,
        }
    })

def export_reports_to_excel(request):
    """Export filtered reports to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    # Get the same filtered reports as the main view
    status_filter = request.GET.get('status', '')
    severity_filter = request.GET.get('severity', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')
    
    reports = IncidentReport.objects.select_related(
        'reporter', 'incident_type', 'reported_student', 'classification', 'sanction'
    ).order_by('-created_at')
    
    # Apply the same filters
    if status_filter:
        reports = reports.filter(status=status_filter)
    if severity_filter and severity_filter != 'all':
        reports = reports.filter(classification__severity=severity_filter)
    if date_from:
        try:
            from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
            reports = reports.filter(incident_date__gte=from_date)
        except ValueError:
            pass
    if date_to:
        try:
            to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
            reports = reports.filter(incident_date__lte=to_date)
        except ValueError:
            pass
    if search_query:
        reports = reports.filter(
            Q(case_id__icontains=search_query) |
            Q(reported_student__first_name__icontains=search_query) |
            Q(reported_student__last_name__icontains=search_query) |
            Q(incident_type__name__icontains=search_query)
        )
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incident Reports"
    
    # Define comprehensive headers
    headers = [
        'Case ID', 'Incident Date', 'Filed Date', 'Student Name', 'Student ID',
        'Reporter Name', 'Reporter Role', 'Incident Type', 'Severity',
        'Classification', 'Status', 'Grade Level', 'Section', 'Teacher',
        'Description', 'Involved Students', 'Sanction Type', 'Sanction Duration (Days)',
        'Counseling Sessions', 'Evaluation Notes', 'Verdict'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data with complete details (no empty cells)
    for row, report in enumerate(reports, 2):
        # Sanction information
        if hasattr(report, 'sanction') and report.sanction:
            sanction_type = report.sanction.get_sanction_type_display()
            sanction_duration = str(report.sanction.duration_days) if report.sanction.duration_days else 'N/A'
            sanction_reason = report.sanction.reason if report.sanction.reason else 'Not specified'
        else:
            sanction_type = 'No Sanction Issued'
            sanction_duration = 'N/A'
            sanction_reason = 'N/A'
        
        # Classification information
        if hasattr(report, 'classification') and report.classification:
            classification_severity = report.classification.get_severity_display()
            severity_type = report.classification.severity.upper()
        else:
            classification_severity = 'Not Yet Classified'
            severity_type = 'Pending'
        
        # Evaluation information
        if hasattr(report, 'evaluation') and report.evaluation:
            evaluation_notes = report.evaluation.evaluation_notes[:100] + '...' if len(report.evaluation.evaluation_notes) > 100 else report.evaluation.evaluation_notes
            verdict = report.evaluation.get_verdict_display() if hasattr(report.evaluation, 'verdict') else 'Not determined'
        else:
            evaluation_notes = 'Not yet evaluated'
            verdict = 'Pending evaluation'
        
        # Counseling sessions count
        counseling_count = report.counseling_sessions.count() if hasattr(report, 'counseling_sessions') else 0
        counseling_info = f"{counseling_count} session(s)" if counseling_count > 0 else 'No sessions scheduled'
        
        # Student information
        student_name = report.reported_student.get_full_name() if report.reported_student else (report.involved_students if report.involved_students else 'Not specified')
        student_id = report.reported_student.username if report.reported_student else 'N/A'
        
        # Reporter information
        reporter_name = report.reporter.get_full_name() if report.reporter else 'Anonymous'
        reporter_role = report.reporter.get_role_display() if report.reporter else 'N/A'
        
        # Incident details
        incident_type = report.incident_type.name if report.incident_type else 'Not specified'
        description = report.description if report.description else 'No description provided'
        involved_students = report.involved_students if report.involved_students else 'Not specified'
        
        # Academic information
        grade_level = f"Grade {report.grade_level}" if report.grade_level else 'Not specified'
        section = report.section_name if report.section_name else 'Not specified'
        teacher = report.teacher_name if report.teacher_name else 'Not specified'
        
        # Write all data to Excel
        ws.cell(row=row, column=1, value=report.case_id or 'N/A')
        ws.cell(row=row, column=2, value=report.incident_date.strftime('%Y-%m-%d') if report.incident_date else 'Not specified')
        ws.cell(row=row, column=3, value=report.created_at.strftime('%Y-%m-%d %H:%M') if report.created_at else 'N/A')
        ws.cell(row=row, column=4, value=student_name)
        ws.cell(row=row, column=5, value=student_id)
        ws.cell(row=row, column=6, value=reporter_name)
        ws.cell(row=row, column=7, value=reporter_role)
        ws.cell(row=row, column=8, value=incident_type)
        ws.cell(row=row, column=9, value=severity_type)
        ws.cell(row=row, column=10, value=classification_severity)
        ws.cell(row=row, column=11, value=report.get_status_display())
        ws.cell(row=row, column=12, value=grade_level)
        ws.cell(row=row, column=13, value=section)
        ws.cell(row=row, column=14, value=teacher)
        ws.cell(row=row, column=15, value=description)
        ws.cell(row=row, column=16, value=involved_students)
        ws.cell(row=row, column=17, value=sanction_type)
        ws.cell(row=row, column=18, value=sanction_duration)
        ws.cell(row=row, column=19, value=sanction_reason)
        ws.cell(row=row, column=20, value=counseling_info)
        ws.cell(row=row, column=21, value=evaluation_notes)
        ws.cell(row=row, column=22, value=verdict)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with current date
    filename = f"incident_reports_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    return response

@login_required
def internal_notes(request):
    if request.user.role != 'do':
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = InternalNoteForm(request.POST)
        if form.is_valid():
            note = form.save(commit=False)
            note.author = request.user
            report_id = request.POST.get('report_id')
            note.report = get_object_or_404(IncidentReport, id=report_id)
            note.save()
            messages.success(request, 'Internal note added successfully')
            return redirect('internal_notes')
    
    notes = InternalNote.objects.select_related('report', 'author').all()
    form = InternalNoteForm()
    reports = IncidentReport.objects.all()
    
    return render(request, 'do/internal_notes.html', {
        'notes': notes,
        'form': form,
        'reports': reports
    })

# Counselor-specific views
@login_required
def major_case_review(request):
    if request.user.role != 'counselor':
        return redirect('dashboard')
    
    # Handle LRN submission
    if request.method == 'POST' and request.POST.get('action') == 'add_lrn':
        case_id = request.POST.get('case_id')
        lrn = request.POST.get('lrn', '').strip()
        
        if case_id and lrn:
            try:
                report = IncidentReport.objects.get(id=case_id)
                # Try to find or create a student with this LRN
                student, created = CustomUser.objects.get_or_create(
                    username=lrn,
                    defaults={
                        'role': 'student',
                        'first_name': 'Student',
                        'last_name': lrn,
                    }
                )
                report.reported_student = student
                report.save()
                messages.success(request, f'LRN {lrn} added to case {report.case_id}')
            except IncidentReport.DoesNotExist:
                messages.error(request, 'Case not found')
            except Exception as e:
                messages.error(request, f'Error adding LRN: {str(e)}')
        
        return redirect('major_case_review')
    
    # Handle Excel export
    if request.GET.get('export') == 'excel':
        return export_counselor_reports_to_excel(request)
    
    # Get all reports (not just major cases)
    all_reports = IncidentReport.objects.all().select_related(
        'classification', 'incident_type', 'reported_student', 'reporter'
    ).order_by('-created_at')
    
    # Apply search filter
    search_query = request.GET.get('search', '').strip()
    if search_query:
        all_reports = all_reports.filter(
            Q(reported_student__first_name__icontains=search_query) |
            Q(reported_student__last_name__icontains=search_query) |
            Q(reported_student__username__icontains=search_query) |
            Q(case_id__icontains=search_query) |
            Q(involved_students__icontains=search_query)
        )
    
    # Calculate statistics
    from django.utils import timezone
    from django.db.models import Count
    today = timezone.now().date()
    
    total_reports = all_reports.count()
    major_cases = all_reports.filter(classification__severity='major').count()
    pending_evaluation = all_reports.filter(classification__severity='major', evaluation__isnull=True).count()
    
    # Find repeat offenders - students with more than 1 violation
    repeat_offenders_data = ViolationHistory.objects.values('student').annotate(
        violation_count=Count('id')
    ).filter(violation_count__gt=1)
    repeat_offenders = repeat_offenders_data.count()
    
    # Get list of repeat offender student IDs for highlighting
    repeat_offender_ids = [item['student'] for item in repeat_offenders_data]
    
    # Count evaluations done today
    evaluated_today = CaseEvaluation.objects.filter(
        evaluated_at__date=today
    ).count()
    
    context = {
        'cases': all_reports,
        'total_reports': total_reports,
        'major_cases': major_cases,
        'pending_evaluation': pending_evaluation,
        'repeat_offenders': repeat_offenders,
        'repeat_offender_ids': repeat_offender_ids,
        'evaluated_today': evaluated_today,
        'search_query': search_query,
    }
    
    return render(request, 'counselor/major_case_review.html', context)

def export_counselor_reports_to_excel(request):
    """Export all reports to Excel for counselors with repeat offender detection"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from django.db.models import Count
    
    # Get all reports
    reports = IncidentReport.objects.all().select_related(
        'reporter', 'incident_type', 'reported_student', 'classification'
    ).order_by('-created_at')
    
    # Apply search filter if provided
    search_query = request.GET.get('search', '').strip()
    if search_query:
        reports = reports.filter(
            Q(reported_student__first_name__icontains=search_query) |
            Q(reported_student__last_name__icontains=search_query) |
            Q(reported_student__username__icontains=search_query) |
            Q(case_id__icontains=search_query)
        )
    
    # Get repeat offenders
    repeat_offenders_data = ViolationHistory.objects.values('student', 'student__first_name', 'student__last_name', 'student__username').annotate(
        violation_count=Count('id')
    ).filter(violation_count__gt=1).order_by('-violation_count')
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: All Reports
    ws1 = wb.active
    ws1.title = "All Reports"
    
    # Headers
    headers = [
        'Case ID', 'Student Name', 'LRN', 'Grade', 'Section',
        'Incident Type', 'Incident Date', 'Reporter', 'Status',
        'Classification', 'Description', 'Repeat Offender'
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Get repeat offender IDs
    repeat_offender_ids = [item['student'] for item in repeat_offenders_data]
    
    # Add data
    for row, report in enumerate(reports, 2):
        student_name = report.reported_student.get_full_name() if report.reported_student else report.involved_students or 'Not specified'
        lrn = report.reported_student.username if report.reported_student else 'N/A'
        is_repeat = 'YES' if report.reported_student and report.reported_student.id in repeat_offender_ids else 'NO'
        
        ws1.cell(row=row, column=1, value=report.case_id or 'N/A')
        ws1.cell(row=row, column=2, value=student_name)
        ws1.cell(row=row, column=3, value=lrn)
        ws1.cell(row=row, column=4, value=f"Grade {report.grade_level}" if report.grade_level else 'N/A')
        ws1.cell(row=row, column=5, value=report.section_name or 'N/A')
        ws1.cell(row=row, column=6, value=report.incident_type.name if report.incident_type else 'Not specified')
        ws1.cell(row=row, column=7, value=report.incident_date.strftime('%Y-%m-%d') if report.incident_date else 'N/A')
        ws1.cell(row=row, column=8, value=report.reporter.get_full_name() if report.reporter else 'Anonymous')
        ws1.cell(row=row, column=9, value=report.get_status_display())
        ws1.cell(row=row, column=10, value=report.classification.get_severity_display() if hasattr(report, 'classification') else 'Not classified')
        ws1.cell(row=row, column=11, value=report.description[:100] if report.description else 'No description')
        
        # Highlight repeat offenders
        repeat_cell = ws1.cell(row=row, column=12, value=is_repeat)
        if is_repeat == 'YES':
            repeat_cell.font = Font(bold=True, color="FF0000")
            repeat_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Repeat Offenders Summary
    ws2 = wb.create_sheet("Repeat Offenders")
    
    headers2 = ['Student Name', 'LRN', 'Total Violations']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row, offender in enumerate(repeat_offenders_data, 2):
        ws2.cell(row=row, column=1, value=f"{offender['student__first_name']} {offender['student__last_name']}")
        ws2.cell(row=row, column=2, value=offender['student__username'])
        ws2.cell(row=row, column=3, value=offender['violation_count'])
    
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"all_reports_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def major_case_detail(request, case_id):
    if request.user.role != 'counselor':
        return redirect('dashboard')
    
    case = get_object_or_404(
        IncidentReport.objects.select_related(
            'classification', 'incident_type', 'reported_student', 'reporter'
        ),
        id=case_id
    )
    
    # Get student violation history
    student_history = ViolationHistory.objects.filter(
        student=case.reported_student
    ).select_related('violation_type').order_by('-date_occurred')
    
    # Calculate days open
    from django.utils import timezone
    days_open = (timezone.now().date() - case.created_at.date()).days
    
    context = {
        'case': case,
        'student_history': student_history,
        'days_open': days_open,
    }
    
    return render(request, 'counselor/major_case_detail.html', context)

@login_required
def counseling_management(request):
    if request.user.role != 'counselor':
        return redirect('dashboard')
    
    sessions = CounselingSession.objects.filter(counselor=request.user).order_by('-scheduled_date')
    
    if request.method == 'POST':
        form = CounselingSessionForm(request.POST)
        if form.is_valid():
            session = form.save(commit=False)
            session.counselor = request.user
            session.report = None  # No specific report linked
            session.save()
            
            # Create notification for the student (violator/victim)
            Notification.objects.create(
                user=session.student,
                title='Counseling Session Scheduled',
                message=f'A counseling session has been scheduled for you on {session.scheduled_date.strftime("%B %d, %Y at %I:%M %p")}. Please be present at the Guidance Office on time.',
                report=None
            )
            
            # Get all reports involving this student
            student_reports = IncidentReport.objects.filter(reported_student=session.student)
            
            # Notify all reporters (teachers/students who reported this student)
            reporters_to_notify = {}  # Use dict to store reporter and their report
            for report in student_reports:
                if report.reporter:
                    reporters_to_notify[report.reporter.id] = {
                        'user': report.reporter,
                        'report': report
                    }
            
            for reporter_data in reporters_to_notify.values():
                Notification.objects.create(
                    user=reporter_data['user'],
                    title='Counseling Session Scheduled',
                    message=f'A counseling session has been scheduled for {session.student.get_full_name()} on {session.scheduled_date.strftime("%B %d, %Y at %I:%M %p")}. This is regarding the incident report you filed (Case ID: {reporter_data["report"].case_id}).',
                    report=reporter_data['report']
                )
            
            # Notify Discipline Officers
            do_users = CustomUser.objects.filter(role='do', is_active=True)
            for do_user in do_users:
                Notification.objects.create(
                    user=do_user,
                    title='Counseling Session Scheduled',
                    message=f'Counseling session scheduled for {session.student.get_full_name()} on {session.scheduled_date.strftime("%B %d, %Y at %I:%M %p")} by {request.user.get_full_name()}.',
                    report=None
                )
            
            notification_count = len(reporters_to_notify) + len(do_users) + 1  # reporters + DOs + student
            messages.success(request, f'Counseling session scheduled successfully for {session.student.get_full_name()}. {notification_count} notification(s) sent.')
            return redirect('counseling_management')
    
    # Calculate session statistics
    session_stats = {
        'total': sessions.count(),
        'completed': sessions.filter(status='completed').count(),
        'scheduled': sessions.filter(status='scheduled').count(),
        'cancelled': sessions.filter(status='cancelled').count(),
    }
    
    # Separate sessions by status
    scheduled_sessions = sessions.filter(status='scheduled')
    completed_sessions = sessions.filter(status='completed')
    
    # Check if a specific student is pre-selected
    selected_student_id = request.GET.get('student_id')
    selected_student = None
    if selected_student_id:
        try:
            selected_student = CustomUser.objects.get(id=selected_student_id, role='student')
        except CustomUser.DoesNotExist:
            pass
    
    # Get reported students with their incident counts for context
    reported_students_info = []
    reported_student_ids = IncidentReport.objects.values_list('reported_student', flat=True).distinct()
    for student_id in reported_student_ids:
        if student_id:  # Make sure student_id is not None
            try:
                student = CustomUser.objects.get(id=student_id, role='student')
                incident_count = IncidentReport.objects.filter(reported_student=student).count()
                latest_incident = IncidentReport.objects.filter(reported_student=student).order_by('-created_at').first()
                reported_students_info.append({
                    'student': student,
                    'incident_count': incident_count,
                    'latest_incident': latest_incident
                })
            except CustomUser.DoesNotExist:
                continue
    
    form = CounselingSessionForm()
    return render(request, 'counselor/counseling_management_clean.html', {
        'sessions': sessions,
        'scheduled_sessions': scheduled_sessions,
        'completed_sessions': completed_sessions,
        'selected_student': selected_student,
        'reported_students_info': reported_students_info,
        'form': form,
        'session_stats': session_stats,
        'today': timezone.now().date(),
        'now': timezone.now()
    })

@login_required
def complete_counseling_session(request, session_id):
    if request.user.role != 'counselor':
        return JsonResponse({'success': False, 'message': 'Unauthorized'})
    
    if request.method == 'POST':
        try:
            session = get_object_or_404(CounselingSession, id=session_id, counselor=request.user)
            session.status = 'completed'
            session.save()
            
            # Create notification for student
            Notification.objects.create(
                user=session.student,
                title='Counseling Session Completed',
                message=f'Your counseling session scheduled for {session.scheduled_date.strftime("%B %d, %Y")} has been marked as completed.',
                report=session.report
            )
            
            return JsonResponse({'success': True, 'message': 'Session marked as completed'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def reschedule_counseling_session(request, session_id):
    if request.user.role != 'counselor':
        return JsonResponse({'success': False, 'message': 'Unauthorized'})
    
    if request.method == 'POST':
        try:
            import json
            from datetime import datetime
            
            session = get_object_or_404(CounselingSession, id=session_id, counselor=request.user)
            data = json.loads(request.body)
            new_date_str = data.get('scheduled_date')
            
            if not new_date_str:
                return JsonResponse({'success': False, 'message': 'New date is required'})
            
            # Parse the new date
            try:
                new_date = datetime.strptime(new_date_str, '%Y-%m-%d %H:%M')
            except ValueError:
                return JsonResponse({'success': False, 'message': 'Invalid date format. Use YYYY-MM-DD HH:MM'})
            
            old_date = session.scheduled_date
            session.scheduled_date = new_date
            session.save()
            
            # Create notification for student
            Notification.objects.create(
                user=session.student,
                title='Counseling Session Rescheduled',
                message=f'Your counseling session has been rescheduled from {old_date.strftime("%B %d, %Y at %I:%M %p")} to {new_date.strftime("%B %d, %Y at %I:%M %p")}.',
                report=session.report
            )
            
            return JsonResponse({'success': True, 'message': 'Session rescheduled successfully'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def cancel_counseling_session(request, session_id):
    if request.user.role != 'counselor':
        return JsonResponse({'success': False, 'message': 'Unauthorized'})
    
    if request.method == 'POST':
        try:
            session = get_object_or_404(CounselingSession, id=session_id, counselor=request.user)
            session.status = 'cancelled'
            session.save()
            
            # Create notification for student
            Notification.objects.create(
                user=session.student,
                title='Counseling Session Cancelled',
                message=f'Your counseling session scheduled for {session.scheduled_date.strftime("%B %d, %Y at %I:%M %p")} has been cancelled.',
                report=session.report
            )
            
            return JsonResponse({'success': True, 'message': 'Session cancelled successfully'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def case_evaluation(request):
    if request.user.role != 'counselor':
        return redirect('dashboard')
    
    if request.method == 'POST':
        report_id = request.POST.get('report_id')
        commission = request.POST.get('commission')
        intervention = request.POST.get('intervention')
        status = request.POST.get('status')
        notes = request.POST.get('notes', '')
        
        try:
            report = IncidentReport.objects.get(id=report_id)
        except IncidentReport.DoesNotExist:
            messages.error(request, 'Report not found.')
            return redirect('case_evaluation')
        
        # Check if report has a student
        if not report.reported_student:
            messages.error(request, f'Cannot evaluate case {report.case_id}: No student associated with this report.')
            return redirect('case_evaluation')
        
        # Check if evaluation already exists
        if hasattr(report, 'evaluation'):
            messages.error(request, f'Case {report.case_id} has already been evaluated.')
            return redirect('case_evaluation')
        
        # Determine recommendation based on intervention
        recommendation = 'counseling'  # default
        if intervention and 'VPF' in intervention:
            recommendation = 'monitoring'
        
        # Create evaluation
        try:
            evaluation = CaseEvaluation.objects.create(
                report=report,
                evaluated_by=request.user,
                recommendation=recommendation,
                verdict='pending',
                is_repeat_offender=False,
                evaluation_notes=f"Commission: {commission}\nIntervention: {intervention}\nStatus: {status}\n{notes}"
            )
            
            # Update report status
            report.status = 'evaluated'
            report.save()
            
            # If VPF selected, create VPF case for ESP Teacher to manage
            if intervention and ('VPF' in intervention or 'Values Reflective Formation' in intervention):
                vpf_case = VPFCase.objects.create(
                    report=report,
                    student=report.reported_student,
                    assigned_by=request.user,
                    commission_level=commission,
                    intervention=intervention,
                    status='pending',
                    notes=f"Commission: {commission}\nIntervention: {intervention}\nEvaluation Notes: {notes}"
                )
                
                # Notify ALL ESP Teachers (they will manage VPF)
                esp_teachers = CustomUser.objects.filter(role='esp_teacher')
                for esp_teacher in esp_teachers:
                    Notification.objects.create(
                        user=esp_teacher,
                        title='New VPF Case Assigned',
                        message=f'Case {report.case_id} has been assigned for VPF - {commission}. Student: {report.reported_student.get_full_name()}. Please schedule the VPF session.',
                        report=report
                    )
                
                # Notify the reporter (adviser/teacher)
                if report.reporter:
                    Notification.objects.create(
                        user=report.reporter,
                        title='Case Evaluated - VPF Assigned',
                        message=f'Case {report.case_id} has been evaluated. The student will undergo Values Reflective Formation (VPF) - {commission}. ESP Teacher will schedule the session.',
                        report=report
                    )
                
                # Notify the student
                if report.reported_student:
                    Notification.objects.create(
                        user=report.reported_student,
                        title='VPF Case Assigned',
                        message=f'You have been assigned to Values Reflective Formation (VPF) for case {report.case_id}. The ESP Teacher will schedule your session.',
                        report=report
                    )
                
                messages.success(request, f'✅ VPF case created for {report.reported_student.get_full_name()}. ESP Teachers have been notified to schedule the session.')
            else:
                # Non-VPF intervention - Create pending counseling schedule entry
                # This will appear in Counseling Schedule sidebar for counselor to schedule
                from datetime import datetime, timedelta
                
                # Create a placeholder counseling schedule (counselor will set actual date/time later)
                counseling_schedule = CounselingSchedule.objects.create(
                    evaluation=evaluation,
                    counselor=request.user,
                    student=report.reported_student,
                    scheduled_date=datetime.now() + timedelta(days=7),  # Default 7 days from now
                    location='Guidance Office',
                    notes=f"Commission: {commission}\nIntervention: {intervention}\nStatus: {status}\n{notes}",
                    status='scheduled'
                )
                
                # Notify the adviser (reporter)
                if report.reporter:
                    Notification.objects.create(
                        user=report.reporter,
                        title='Case Evaluated - Counseling Scheduled',
                        message=f'Case {report.case_id} has been evaluated. Intervention: {intervention}. The student will be scheduled for counseling.',
                        report=report
                    )
                
                # Notify the student
                if report.reported_student:
                    Notification.objects.create(
                        user=report.reported_student,
                        title='Counseling Session Scheduled',
                        message=f'Your case {report.case_id} has been evaluated. A counseling session will be scheduled. You will receive the schedule details soon.',
                        report=report
                    )
                
                messages.success(request, f'✅ Evaluation completed for {report.reported_student.get_full_name()}. Counseling schedule created. Please set the final schedule in Counseling Schedule.')
            
            return redirect('case_evaluation')
            
        except Exception as e:
            messages.error(request, f'Error creating evaluation: {str(e)}')
            return redirect('case_evaluation')
    
    # Get cases ready for evaluation (not yet evaluated)
    cases_for_evaluation = IncidentReport.objects.filter(
        evaluation__isnull=True,
        reported_student__isnull=False
    ).exclude(
        status__in=['closed', 'resolved']
    ).order_by('-created_at')
    
    return render(request, 'counselor/case_evaluation.html', {
        'cases': cases_for_evaluation,
    })

@login_required
def counselor_schedule(request):
    """Counseling schedule page for non-VPF interventions"""
    if request.user.role != 'counselor':
        return redirect('dashboard')
    
    from .models import CounselingSchedule
    
    if request.method == 'POST':
        evaluation_id = request.POST.get('evaluation_id')
        scheduled_date_str = request.POST.get('scheduled_date')
        location = request.POST.get('location', '')
        notes = request.POST.get('notes', '')
        
        evaluation = get_object_or_404(CaseEvaluation, id=evaluation_id)
        
        # Parse the scheduled date and make it timezone-aware
        from datetime import datetime
        scheduled_date = datetime.fromisoformat(scheduled_date_str)
        
        # Make the datetime timezone-aware if it's naive
        if scheduled_date.tzinfo is None:
            scheduled_date = timezone.make_aware(scheduled_date)
        
        # Validation 1: Check if weekend
        if scheduled_date.weekday() >= 5:  # 5=Saturday, 6=Sunday
            messages.error(request, 'Weekend scheduling is not allowed. Please select a weekday.')
            return redirect(f'/counselor-schedule/?evaluation_id={evaluation_id}')
        
        # Validation 2: Check if in the past
        if scheduled_date < timezone.now():
            messages.error(request, 'Cannot schedule in the past. Please select a future date and time.')
            return redirect(f'/counselor-schedule/?evaluation_id={evaluation_id}')
        
        # Validation 3: Check for duplicate/conflict for this student
        from datetime import timedelta
        time_window_start = scheduled_date - timedelta(hours=1)
        time_window_end = scheduled_date + timedelta(hours=1)
        
        existing_schedule = CounselingSchedule.objects.filter(
            student=evaluation.report.reported_student,
            scheduled_date__range=(time_window_start, time_window_end),
            status__in=['scheduled', 'rescheduled']
        ).exists()
        
        if existing_schedule:
            messages.error(request, f'Schedule conflict: {evaluation.report.reported_student.get_full_name()} already has a session scheduled within this time window. Please choose a different time.')
            return redirect(f'/counselor-schedule/?evaluation_id={evaluation_id}')
        
        # Create counseling schedule
        schedule = CounselingSchedule.objects.create(
            evaluation=evaluation,
            counselor=request.user,
            student=evaluation.report.reported_student,
            scheduled_date=scheduled_date,
            location=location,
            notes=notes,
            status='scheduled'
        )
        
        # Notify the student
        Notification.objects.create(
            user=evaluation.report.reported_student,
            title='Counseling Session Scheduled',
            message=f'Your counseling session for case {evaluation.report.case_id} has been scheduled for {schedule.scheduled_date.strftime("%B %d, %Y at %I:%M %p")}. Location: {location if location else "TBA"}',
            report=evaluation.report
        )
        
        # Notify the reporter
        if evaluation.report.reporter:
            Notification.objects.create(
                user=evaluation.report.reporter,
                title='Counseling Session Scheduled',
                message=f'A counseling session for case {evaluation.report.case_id} has been scheduled for {schedule.scheduled_date.strftime("%B %d, %Y at %I:%M %p")}.',
                report=evaluation.report
            )
        
        # Notify the adviser/teacher based on student's curriculum, grade, and section
        notify_adviser_of_counseling(
            report=evaluation.report,
            scheduled_date=scheduled_date,
            location=location if location else "TBA",
            counselor_name=request.user.get_full_name(),
            schedule_type='Counseling'
        )
        
        messages.success(request, f'Counseling session scheduled for {evaluation.report.reported_student.get_full_name()}. Notifications sent.')
        return redirect('counselor_schedule')
    
    # Get pending evaluations that need scheduling (non-VPF)
    pending_evaluations = CaseEvaluation.objects.filter(
        evaluated_by=request.user,
        counseling_schedules__isnull=True
    ).exclude(
        report__vpf_cases__isnull=False
    ).select_related('report', 'report__reported_student')
    
    # Get all scheduled sessions
    schedules = CounselingSchedule.objects.filter(
        counselor=request.user
    ).select_related('evaluation__report', 'student').order_by('scheduled_date')
    
    # Check if a specific evaluation is pre-selected
    selected_evaluation_id = request.GET.get('evaluation_id')
    selected_evaluation = None
    if selected_evaluation_id:
        try:
            selected_evaluation = pending_evaluations.get(id=selected_evaluation_id)
        except CaseEvaluation.DoesNotExist:
            pass
    
    return render(request, 'counselor/counselor_schedule.html', {
        'pending_evaluations': pending_evaluations,
        'schedules': schedules,
        'selected_evaluation': selected_evaluation,
        'today': timezone.now().date()
    })

@login_required
def complete_counseling_schedule(request, schedule_id):
    """Mark a counseling schedule as completed"""
    if request.user.role != 'counselor':
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    from .models import CounselingSchedule
    
    schedule = get_object_or_404(CounselingSchedule, id=schedule_id, counselor=request.user)
    schedule.status = 'completed'
    schedule.save()
    
    # Notify student
    Notification.objects.create(
        user=schedule.student,
        title='Counseling Session Completed',
        message=f'Your counseling session for case {schedule.evaluation.report.case_id} has been completed.',
        report=schedule.evaluation.report
    )
    
    return JsonResponse({'success': True, 'message': 'Session marked as completed'})

@login_required
def missed_counseling_schedule(request, schedule_id):
    """Mark a counseling schedule as missed"""
    if request.user.role != 'counselor':
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    from .models import CounselingSchedule
    
    schedule = get_object_or_404(CounselingSchedule, id=schedule_id, counselor=request.user)
    schedule.status = 'missed'
    schedule.save()
    
    # Notify student
    Notification.objects.create(
        user=schedule.student,
        title='Counseling Session Missed',
        message=f'You missed your counseling session for case {schedule.evaluation.report.case_id}. Please contact the guidance office to reschedule.',
        report=schedule.evaluation.report
    )
    
    return JsonResponse({'success': True, 'message': 'Session marked as missed'})

# Principal-specific views
@login_required
def evaluated_cases(request):
    if request.user.role != 'principal':
        return redirect('dashboard')
    
    evaluated_cases = IncidentReport.objects.filter(
        status='evaluated'
    ).select_related('evaluation', 'classification')
    
    return render(request, 'principal/evaluated_cases.html', {'cases': evaluated_cases})

@login_required
def sanction_management(request):
    if request.user.role != 'principal':
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = SanctionForm(request.POST)
        if form.is_valid():
            report_id = request.POST.get('report_id')
            report = get_object_or_404(IncidentReport, id=report_id)
            
            # Check if sanction already exists for this report
            try:
                existing_sanction = report.sanction
                messages.error(request, f'A sanction has already been issued for case {report.case_id}.')
                return redirect('sanction_management')
            except:
                pass  # No sanction exists, continue
            
            try:
                sanction = form.save(commit=False)
                sanction.issued_by = request.user
                sanction.report = report
                sanction.save()
                
                # Update report status
                report.status = 'sanctioned'
                report.save()
                
                # Send notifications
                Notification.objects.create(
                    user=report.reported_student,
                    title='Sanction Issued',
                    message=f'A {sanction.get_sanction_type_display()} has been issued for case {report.case_id}.',
                    report=report
                )
                
                # Notify guidance for post-counseling
                counselors = CustomUser.objects.filter(role='counselor')
                for counselor in counselors:
                    Notification.objects.create(
                        user=counselor,
                        title='Post-Sanction Counseling Required',
                        message=f'Post-sanction counseling required for case {report.case_id}.',
                        report=report
                    )
                
                messages.success(request, f'Sanction issued successfully for case {report.case_id}')
                return redirect('sanction_management')
                
            except Exception as e:
                messages.error(request, f'Error issuing sanction: A sanction may already exist for this case.')
                return redirect('sanction_management')
    
    # Filter out cases that already have sanctions
    cases_for_sanction = IncidentReport.objects.filter(
        status='evaluated',
        sanction__isnull=True  # Only cases without existing sanctions
    ).select_related('evaluation', 'incident_type', 'reported_student')
    
    form = SanctionForm()
    
    # Get recently issued sanctions for display
    recent_sanctions = Sanction.objects.select_related(
        'report', 'report__reported_student'
    ).order_by('-issued_at')[:5]
    
    # Calculate statistics
    sanction_stats = {
        'pending': cases_for_sanction.count(),
        'total_sanctions': Sanction.objects.count(),
        'warnings': Sanction.objects.filter(sanction_type='warning').count(),
        'suspensions': Sanction.objects.filter(sanction_type='suspension').count(),
        'expulsions': Sanction.objects.filter(sanction_type='expulsion').count(),
    }
    
    return render(request, 'principal/sanction_management.html', {
        'cases': cases_for_sanction,
        'recent_sanctions': recent_sanctions,
        'sanction_stats': sanction_stats,
        'form': form
    })

@login_required
def final_verdicts(request):
    if request.user.role != 'principal':
        return redirect('dashboard')
    
    sanctioned_cases = IncidentReport.objects.filter(
        status='sanctioned'
    ).select_related('sanction').order_by('-sanction__issued_at')
    
    return render(request, 'principal/final_verdicts.html', {'cases': sanctioned_cases})

@login_required
def student_monitoring(request):
    if request.user.role != 'principal':
        return redirect('dashboard')
    
    # Get all reports with related data
    all_reports = IncidentReport.objects.select_related(
        'reported_student', 'incident_type', 'reporter'
    ).prefetch_related('counseling_sessions').order_by('-incident_date', '-incident_time')
    
    # Filter by status if requested
    status_filter = request.GET.get('status')
    if status_filter:
        all_reports = all_reports.filter(status=status_filter)
    
    # Filter by grade if requested
    grade_filter = request.GET.get('grade')
    if grade_filter:
        all_reports = all_reports.filter(grade_level=grade_filter)
    
    # Get violation trends
    violation_trends = ViolationHistory.objects.values('student__grade_level').annotate(
        violation_count=models.Count('id')
    ).order_by('-violation_count')
    
    # Calculate statistics
    total_students = CustomUser.objects.filter(role='student').count()
    students_with_violations = IncidentReport.objects.values('reported_student').distinct().count()
    repeat_offenders = ViolationHistory.objects.values('student').annotate(
        violation_count=models.Count('id')
    ).filter(violation_count__gt=1).count()
    
    # Get unique grade levels for filter
    grade_levels = IncidentReport.objects.values_list('grade_level', flat=True).distinct().order_by('grade_level')
    
    # Status choices for filter
    status_choices = IncidentReport.STATUS_CHOICES
    
    monitoring_stats = {
        'total_students': total_students,
        'students_with_violations': students_with_violations,
        'repeat_offenders': repeat_offenders,
        'under_monitoring': all_reports.filter(status__in=['classified', 'evaluated']).count(),
        'total_reports': all_reports.count(),
    }
    
    return render(request, 'principal/student_monitoring.html', {
        'all_reports': all_reports,
        'violation_trends': violation_trends,
        'monitoring_stats': monitoring_stats,
        'grade_levels': grade_levels,
        'status_choices': status_choices,
        'current_status': status_filter,
        'current_grade': grade_filter,
    })

# Shared views
@login_required
def reports_analytics(request):
    if request.user.role not in ['do', 'counselor', 'principal']:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = ReportAnalyticsForm(request.POST)
        if form.is_valid():
            # Generate analytics based on form data
            start_date = form.cleaned_data['date_range_start']
            end_date = form.cleaned_data['date_range_end']
            report_type = form.cleaned_data['report_type']
            
            # Get reports in date range
            reports = IncidentReport.objects.filter(
                created_at__date__range=[start_date, end_date]
            )
            
            analytics_data = {
                'total_reports': reports.count(),
                'minor_violations': reports.filter(classification__severity='minor').count(),
                'major_violations': reports.filter(classification__severity='major').count(),
                'resolved_cases': reports.filter(status='resolved').count(),
                'pending_cases': reports.filter(status='pending').count(),
                'reports': reports,
                'date_range': f"{start_date} to {end_date}",
                'report_type': report_type
            }
            
            return render(request, 'shared/analytics_results.html', analytics_data)
    
    form = ReportAnalyticsForm()
    return render(request, 'shared/reports_analytics.html', {'form': form})

# Maintenance views (Counselor only)
@login_required
def manage_curriculum(request):
    if request.user.role not in ['guidance', 'maintenance', 'counselor']:
        return redirect('dashboard')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Curriculum CRUD operations
        if action == 'add_curriculum':
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            
            if name:
                if not Curriculum.objects.filter(name=name).exists():
                    Curriculum.objects.create(name=name, description=description)
                    messages.success(request, f'Curriculum "{name}" added successfully.')
                else:
                    messages.error(request, f'Curriculum "{name}" already exists.')
            else:
                messages.error(request, 'Curriculum name is required.')
                
        elif action == 'edit_curriculum':
            curriculum_id = request.POST.get('curriculum_id')
            name = request.POST.get('name', '').strip()
            description = request.POST.get('description', '').strip()
            
            if curriculum_id and name:
                try:
                    curriculum = Curriculum.objects.get(id=curriculum_id)
                    if not Curriculum.objects.filter(name=name).exclude(id=curriculum_id).exists():
                        curriculum.name = name
                        curriculum.description = description
                        curriculum.save()
                        messages.success(request, f'Curriculum "{name}" updated successfully.')
                    else:
                        messages.error(request, f'Curriculum "{name}" already exists.')
                except Curriculum.DoesNotExist:
                    messages.error(request, 'Curriculum not found.')
            else:
                messages.error(request, 'Invalid curriculum data.')
                
        elif action == 'delete_curriculum':
            curriculum_id = request.POST.get('curriculum_id')
            if curriculum_id:
                try:
                    curriculum = Curriculum.objects.get(id=curriculum_id)
                    curriculum_name = curriculum.name
                    curriculum.delete()
                    messages.success(request, f'Curriculum "{curriculum_name}" deleted successfully.')
                except Curriculum.DoesNotExist:
                    messages.error(request, 'Curriculum not found.')
                    
        # Track CRUD operations
        elif action == 'add_track':
            name = request.POST.get('name', '').strip()
            curriculum_id = request.POST.get('curriculum')
            description = request.POST.get('description', '').strip()
            
            if name and curriculum_id:
                try:
                    curriculum = Curriculum.objects.get(id=curriculum_id)
                    if not Track.objects.filter(name=name, curriculum=curriculum).exists():
                        Track.objects.create(name=name, curriculum=curriculum, description=description)
                        messages.success(request, f'Track "{name}" added successfully.')
                    else:
                        messages.error(request, f'Track "{name}" already exists in this curriculum.')
                except Curriculum.DoesNotExist:
                    messages.error(request, 'Selected curriculum not found.')
            else:
                messages.error(request, 'Track name and curriculum are required.')
                
        elif action == 'edit_track':
            track_id = request.POST.get('track_id')
            name = request.POST.get('name', '').strip()
            curriculum_id = request.POST.get('curriculum')
            description = request.POST.get('description', '').strip()
            
            if track_id and name and curriculum_id:
                try:
                    track = Track.objects.get(id=track_id)
                    curriculum = Curriculum.objects.get(id=curriculum_id)
                    if not Track.objects.filter(name=name, curriculum=curriculum).exclude(id=track_id).exists():
                        track.name = name
                        track.curriculum = curriculum
                        track.description = description
                        track.save()
                        messages.success(request, f'Track "{name}" updated successfully.')
                    else:
                        messages.error(request, f'Track "{name}" already exists in this curriculum.')
                except (Track.DoesNotExist, Curriculum.DoesNotExist):
                    messages.error(request, 'Track or curriculum not found.')
            else:
                messages.error(request, 'Invalid track data.')
                
        elif action == 'delete_track':
            track_id = request.POST.get('track_id')
            if track_id:
                try:
                    track = Track.objects.get(id=track_id)
                    track_name = track.name
                    track.delete()
                    messages.success(request, f'Track "{track_name}" deleted successfully.')
                except Track.DoesNotExist:
                    messages.error(request, 'Track not found.')
                    
        # Grade CRUD operations
        elif action == 'add_grade':
            level = request.POST.get('level', '').strip()
            track_id = request.POST.get('track')
            description = request.POST.get('description', '').strip()
            
            if level and track_id:
                try:
                    track = Track.objects.get(id=track_id)
                    if not Grade.objects.filter(level=level, track=track).exists():
                        Grade.objects.create(level=level, track=track, description=description)
                        messages.success(request, f'Grade level "{level}" added successfully.')
                    else:
                        messages.error(request, f'Grade level "{level}" already exists in this track.')
                except Track.DoesNotExist:
                    messages.error(request, 'Selected track not found.')
            else:
                messages.error(request, 'Grade level and track are required.')
                
        elif action == 'edit_grade':
            grade_id = request.POST.get('grade_id')
            level = request.POST.get('level', '').strip()
            track_id = request.POST.get('track')
            description = request.POST.get('description', '').strip()
            
            if grade_id and level and track_id:
                try:
                    grade = Grade.objects.get(id=grade_id)
                    track = Track.objects.get(id=track_id)
                    if not Grade.objects.filter(level=level, track=track).exclude(id=grade_id).exists():
                        grade.level = level
                        grade.track = track
                        grade.description = description
                        grade.save()
                        messages.success(request, f'Grade level "{level}" updated successfully.')
                    else:
                        messages.error(request, f'Grade level "{level}" already exists in this track.')
                except (Grade.DoesNotExist, Track.DoesNotExist):
                    messages.error(request, 'Grade level or track not found.')
            else:
                messages.error(request, 'Invalid grade level data.')
                
        elif action == 'delete_grade':
            grade_id = request.POST.get('grade_id')
            if grade_id:
                try:
                    grade = Grade.objects.get(id=grade_id)
                    grade_level = grade.level
                    grade.delete()
                    messages.success(request, f'Grade level "{grade_level}" deleted successfully.')
                except Grade.DoesNotExist:
                    messages.error(request, 'Grade level not found.')
                    
        # Section CRUD operations
        elif action == 'add_section':
            name = request.POST.get('name', '').strip()
            grade_id = request.POST.get('grade')
            adviser_id = request.POST.get('adviser')
            
            if name and grade_id:
                try:
                    grade = Grade.objects.get(id=grade_id)
                    adviser = CustomUser.objects.get(id=adviser_id) if adviser_id else None
                    if not Section.objects.filter(name=name, grade=grade).exists():
                        Section.objects.create(name=name, grade=grade, adviser=adviser)
                        messages.success(request, f'Section "{name}" added successfully.')
                    else:
                        messages.error(request, f'Section "{name}" already exists in this grade level.')
                except Grade.DoesNotExist:
                    messages.error(request, 'Selected grade level not found.')
                except CustomUser.DoesNotExist:
                    messages.error(request, 'Selected adviser not found.')
            else:
                messages.error(request, 'Section name and grade level are required.')
                
        elif action == 'edit_section':
            section_id = request.POST.get('section_id')
            name = request.POST.get('name', '').strip()
            grade_id = request.POST.get('grade')
            adviser_id = request.POST.get('adviser')
            
            if section_id and name and grade_id:
                try:
                    section = Section.objects.get(id=section_id)
                    grade = Grade.objects.get(id=grade_id)
                    adviser = CustomUser.objects.get(id=adviser_id) if adviser_id else None
                    if not Section.objects.filter(name=name, grade=grade).exclude(id=section_id).exists():
                        section.name = name
                        section.grade = grade
                        section.adviser = adviser
                        section.save()
                        messages.success(request, f'Section "{name}" updated successfully.')
                    else:
                        messages.error(request, f'Section "{name}" already exists in this grade level.')
                except (Section.DoesNotExist, Grade.DoesNotExist):
                    messages.error(request, 'Section or grade level not found.')
                except CustomUser.DoesNotExist:
                    messages.error(request, 'Selected adviser not found.')
            else:
                messages.error(request, 'Invalid section data.')
                
        elif action == 'delete_section':
            section_id = request.POST.get('section_id')
            if section_id:
                try:
                    section = Section.objects.get(id=section_id)
                    section_name = section.name
                    section.delete()
                    messages.success(request, f'Section "{section_name}" deleted successfully.')
                except Section.DoesNotExist:
                    messages.error(request, 'Section not found.')
        
        return redirect('manage_curriculum')
    
    # GET request - display the management page
    curricula = Curriculum.objects.all()
    tracks = Track.objects.select_related('curriculum').all()
    grades = Grade.objects.select_related('track__curriculum').all()
    sections = Section.objects.select_related('grade__track__curriculum', 'adviser').all()
    teachers = CustomUser.objects.filter(role='teacher', is_active=True)
    
    return render(request, 'maintenance/manage_curriculum.html', {
        'curricula': curricula,
        'tracks': tracks,
        'grades': grades,
        'sections': sections,
        'teachers': teachers,
    })

@login_required
def manage_teachers(request):
    if request.user.role not in ['guidance', 'maintenance', 'counselor']:
        return redirect('dashboard')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Teacher CRUD operations
        if action == 'add_teacher':
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '').strip()
            
            if first_name and last_name and email and username and password:
                if not CustomUser.objects.filter(username=username).exists():
                    if not CustomUser.objects.filter(email=email).exists():
                        teacher = CustomUser.objects.create_user(
                            username=username,
                            email=email,
                            password=password,
                            first_name=first_name,
                            last_name=last_name,
                            role='teacher'
                        )
                        messages.success(request, f'Teacher "{teacher.get_full_name()}" added successfully.')
                    else:
                        messages.error(request, f'Email "{email}" is already in use.')
                else:
                    messages.error(request, f'Username "{username}" is already taken.')
            else:
                messages.error(request, 'All fields are required.')
                
        elif action == 'edit_teacher':
            teacher_id = request.POST.get('teacher_id')
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            
            if teacher_id and first_name and last_name and email:
                try:
                    teacher = CustomUser.objects.get(id=teacher_id, role='teacher')
                    if not CustomUser.objects.filter(email=email).exclude(id=teacher_id).exists():
                        teacher.first_name = first_name
                        teacher.last_name = last_name
                        teacher.email = email
                        teacher.save()
                        messages.success(request, f'Teacher "{teacher.get_full_name()}" updated successfully.')
                    else:
                        messages.error(request, f'Email "{email}" is already in use.')
                except CustomUser.DoesNotExist:
                    messages.error(request, 'Teacher not found.')
            else:
                messages.error(request, 'Invalid teacher data.')
                
        elif action == 'toggle_teacher':
            teacher_id = request.POST.get('teacher_id')
            if teacher_id:
                try:
                    teacher = CustomUser.objects.get(id=teacher_id, role='teacher')
                    teacher.is_active = not teacher.is_active
                    teacher.save()
                    status = 'activated' if teacher.is_active else 'deactivated'
                    messages.success(request, f'Teacher "{teacher.get_full_name()}" {status} successfully.')
                except CustomUser.DoesNotExist:
                    messages.error(request, 'Teacher not found.')
                    
        elif action == 'delete_teacher':
            teacher_id = request.POST.get('teacher_id')
            if teacher_id:
                try:
                    teacher = CustomUser.objects.get(id=teacher_id, role='teacher')
                    teacher_name = teacher.get_full_name()
                    teacher.delete()
                    messages.success(request, f'Teacher "{teacher_name}" deleted successfully.')
                except CustomUser.DoesNotExist:
                    messages.error(request, 'Teacher not found.')
                    
        # Teacher Assignment operations
        elif action == 'add_assignment':
            teacher_name = request.POST.get('teacher_name', '').strip()
            curriculum_id = request.POST.get('curriculum_id', '').strip()
            track_code = request.POST.get('track_code', '').strip()
            grade_level = request.POST.get('grade_level', '').strip()
            section_name = request.POST.get('section_name', '').strip()
            
            if teacher_name and curriculum_id and track_code and grade_level and section_name:
                try:
                    curriculum = Curriculum.objects.get(id=curriculum_id)
                    if not TeacherAssignment.objects.filter(
                        teacher_name=teacher_name,
                        curriculum=curriculum,
                        track_code=track_code,
                        grade_level=grade_level, 
                        section_name=section_name
                    ).exists():
                        TeacherAssignment.objects.create(
                            teacher_name=teacher_name,
                            curriculum=curriculum,
                            track_code=track_code,
                            grade_level=grade_level,
                            section_name=section_name
                        )
                        messages.success(request, f'Assignment added for {teacher_name}.')
                    else:
                        messages.error(request, 'This assignment already exists.')
                except Curriculum.DoesNotExist:
                    messages.error(request, 'Curriculum not found.')
            else:
                messages.error(request, 'All assignment fields are required.')
                
        elif action == 'edit_assignment':
            assignment_id = request.POST.get('assignment_id')
            teacher_name = request.POST.get('teacher_name', '').strip()
            curriculum_id = request.POST.get('curriculum_id', '').strip()
            track_code = request.POST.get('track_code', '').strip()
            grade_level = request.POST.get('grade_level', '').strip()
            section_name = request.POST.get('section_name', '').strip()
            
            if assignment_id and teacher_name and curriculum_id and track_code and grade_level and section_name:
                try:
                    assignment = TeacherAssignment.objects.get(id=assignment_id)
                    curriculum = Curriculum.objects.get(id=curriculum_id)
                    
                    # Check if another assignment with same details exists
                    duplicate = TeacherAssignment.objects.filter(
                        teacher_name=teacher_name,
                        curriculum=curriculum,
                        track_code=track_code,
                        grade_level=grade_level,
                        section_name=section_name
                    ).exclude(id=assignment_id).exists()
                    
                    if not duplicate:
                        assignment.teacher_name = teacher_name
                        assignment.curriculum = curriculum
                        assignment.track_code = track_code
                        assignment.grade_level = grade_level
                        assignment.section_name = section_name
                        assignment.save()
                        messages.success(request, 'Assignment updated successfully.')
                    else:
                        messages.error(request, 'This assignment already exists.')
                except TeacherAssignment.DoesNotExist:
                    messages.error(request, 'Assignment not found.')
                except Curriculum.DoesNotExist:
                    messages.error(request, 'Curriculum not found.')
            else:
                messages.error(request, 'All assignment fields are required.')
                
        elif action == 'delete_assignment':
            assignment_id = request.POST.get('assignment_id')
            if assignment_id:
                try:
                    assignment = TeacherAssignment.objects.get(id=assignment_id)
                    assignment.delete()
                    messages.success(request, 'Assignment deleted successfully.')
                except TeacherAssignment.DoesNotExist:
                    messages.error(request, 'Assignment not found.')
        
        return redirect('manage_teachers')
    
    # GET request - display the management page
    teachers = CustomUser.objects.filter(role='teacher').order_by('last_name', 'first_name')
    teacher_assignments = TeacherAssignment.objects.all().select_related('curriculum').order_by('grade_level', 'section_name')
    curricula = Curriculum.objects.all()
    
    return render(request, 'maintenance/manage_teachers.html', {
        'teachers': teachers,
        'teacher_assignments': teacher_assignments,
        'curricula': curricula,
    })

@login_required
def manage_incident_types(request):
    if request.user.role not in ['guidance', 'maintenance', 'counselor']:
        return redirect('dashboard')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Incident Type CRUD operations
        if action == 'add_incident_type':
            name = request.POST.get('name', '').strip()
            severity = request.POST.get('severity', '').strip()
            description = request.POST.get('description', '').strip()
            legal_references = request.POST.get('legal_references', '').strip()
            
            if name and severity:
                if not IncidentType.objects.filter(name=name).exists():
                    IncidentType.objects.create(
                        name=name,
                        severity=severity,
                        description=description,
                        legal_references=legal_references
                    )
                    messages.success(request, f'Incident type "{name}" added successfully.')
                else:
                    messages.error(request, f'Incident type "{name}" already exists.')
            else:
                messages.error(request, 'Incident type name and category are required.')
                
        elif action == 'edit_incident_type':
            incident_type_id = request.POST.get('incident_type_id')
            name = request.POST.get('name', '').strip()
            severity = request.POST.get('severity', '').strip()
            description = request.POST.get('description', '').strip()
            legal_references = request.POST.get('legal_references', '').strip()
            
            if incident_type_id and name and severity:
                try:
                    incident_type = IncidentType.objects.get(id=incident_type_id)
                    if not IncidentType.objects.filter(name=name).exclude(id=incident_type_id).exists():
                        incident_type.name = name
                        incident_type.severity = severity
                        incident_type.description = description
                        incident_type.legal_references = legal_references
                        incident_type.save()
                        messages.success(request, f'Incident type "{name}" updated successfully.')
                    else:
                        messages.error(request, f'Incident type "{name}" already exists.')
                except IncidentType.DoesNotExist:
                    messages.error(request, 'Incident type not found.')
            else:
                messages.error(request, 'Invalid incident type data.')
                
        elif action == 'delete_incident_type':
            incident_type_id = request.POST.get('incident_type_id')
            if incident_type_id:
                try:
                    incident_type = IncidentType.objects.get(id=incident_type_id)
                    incident_type_name = incident_type.name
                    # Check if there are any reports using this incident type
                    reports_count = IncidentReport.objects.filter(incident_type=incident_type).count()
                    if reports_count > 0:
                        messages.warning(request, f'Cannot delete "{incident_type_name}" as it is referenced by {reports_count} incident report(s). Consider deactivating it instead.')
                    else:
                        incident_type.delete()
                        messages.success(request, f'Incident type "{incident_type_name}" deleted successfully.')
                except IncidentType.DoesNotExist:
                    messages.error(request, 'Incident type not found.')
        
        return redirect('manage_incident_types')
    
    # GET request - display the management page
    incident_types = IncidentType.objects.all().order_by('severity', 'name')
    grave_count = incident_types.filter(severity='grave').count()
    prohibited_count = incident_types.filter(severity='prohibited').count()
    
    return render(request, 'maintenance/manage_incident_types.html', {
        'incident_types': incident_types,
        'grave_count': grave_count,
        'prohibited_count': prohibited_count,
    })

@login_required
def manage_legal_references(request):
    if request.user.role not in ['guidance', 'maintenance', 'counselor']:
        return redirect('dashboard')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Legal Reference CRUD operations
        if action == 'add_legal_reference':
            title = request.POST.get('title', '').strip()
            reference_number = request.POST.get('reference_number', '').strip()
            description = request.POST.get('description', '').strip()
            incident_types_ids = request.POST.getlist('incident_types')
            
            if title and reference_number:
                if not LegalReference.objects.filter(reference_number=reference_number).exists():
                    legal_ref = LegalReference.objects.create(
                        title=title,
                        reference_number=reference_number,
                        description=description
                    )
                    if incident_types_ids:
                        legal_ref.incident_types.set(incident_types_ids)
                    messages.success(request, f'Legal reference "{reference_number}" added successfully.')
                else:
                    messages.error(request, f'Legal reference "{reference_number}" already exists.')
            else:
                messages.error(request, 'Title and reference number are required.')
                
        elif action == 'edit_legal_reference':
            legal_ref_id = request.POST.get('legal_reference_id')
            title = request.POST.get('title', '').strip()
            reference_number = request.POST.get('reference_number', '').strip()
            description = request.POST.get('description', '').strip()
            incident_types_ids = request.POST.getlist('incident_types')
            
            if legal_ref_id and title and reference_number:
                try:
                    legal_ref = LegalReference.objects.get(id=legal_ref_id)
                    if not LegalReference.objects.filter(reference_number=reference_number).exclude(id=legal_ref_id).exists():
                        legal_ref.title = title
                        legal_ref.reference_number = reference_number
                        legal_ref.description = description
                        legal_ref.save()
                        if incident_types_ids:
                            legal_ref.incident_types.set(incident_types_ids)
                        else:
                            legal_ref.incident_types.clear()
                        messages.success(request, f'Legal reference "{reference_number}" updated successfully.')
                    else:
                        messages.error(request, f'Legal reference "{reference_number}" already exists.')
                except LegalReference.DoesNotExist:
                    messages.error(request, 'Legal reference not found.')
            else:
                messages.error(request, 'Invalid legal reference data.')
                
        elif action == 'delete_legal_reference':
            legal_ref_id = request.POST.get('legal_reference_id')
            if legal_ref_id:
                try:
                    legal_ref = LegalReference.objects.get(id=legal_ref_id)
                    legal_ref_number = legal_ref.reference_number
                    legal_ref.delete()
                    messages.success(request, f'Legal reference "{legal_ref_number}" deleted successfully.')
                except LegalReference.DoesNotExist:
                    messages.error(request, 'Legal reference not found.')
        
        return redirect('manage_legal_references')
    
    # GET request - display the management page
    legal_references = LegalReference.objects.all().prefetch_related('incident_types')
    incident_types = IncidentType.objects.all()
    
    return render(request, 'maintenance/manage_legal_references.html', {
        'legal_references': legal_references,
        'incident_types': incident_types,
    })

@login_required
def manage_counselors(request):
    if request.user.role not in ['guidance', 'maintenance', 'counselor']:
        return redirect('dashboard')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Counselor CRUD operations
        if action == 'add_counselor':
            name = request.POST.get('name', '').strip()
            email = request.POST.get('email', '').strip()
            phone = request.POST.get('phone', '').strip()
            specialization = request.POST.get('specialization', '').strip()
            
            if name:
                from .models import Counselor
                if not Counselor.objects.filter(name=name).exists():
                    Counselor.objects.create(
                        name=name,
                        email=email,
                        phone=phone,
                        specialization=specialization,
                        is_active=True
                    )
                    messages.success(request, f'Counselor "{name}" added successfully.')
                else:
                    messages.error(request, f'Counselor "{name}" already exists.')
            else:
                messages.error(request, 'Counselor name is required.')
                
        elif action == 'edit_counselor':
            counselor_id = request.POST.get('counselor_id')
            name = request.POST.get('name', '').strip()
            email = request.POST.get('email', '').strip()
            phone = request.POST.get('phone', '').strip()
            specialization = request.POST.get('specialization', '').strip()
            is_active = request.POST.get('is_active') == 'on'
            
            if counselor_id and name:
                try:
                    from .models import Counselor
                    counselor = Counselor.objects.get(id=counselor_id)
                    if not Counselor.objects.filter(name=name).exclude(id=counselor_id).exists():
                        counselor.name = name
                        counselor.email = email
                        counselor.phone = phone
                        counselor.specialization = specialization
                        counselor.is_active = is_active
                        counselor.save()
                        messages.success(request, f'Counselor "{name}" updated successfully.')
                    else:
                        messages.error(request, f'Counselor "{name}" already exists.')
                except Counselor.DoesNotExist:
                    messages.error(request, 'Counselor not found.')
            else:
                messages.error(request, 'Invalid counselor data.')
                
        elif action == 'delete_counselor':
            counselor_id = request.POST.get('counselor_id')
            if counselor_id:
                try:
                    from .models import Counselor
                    counselor = Counselor.objects.get(id=counselor_id)
                    counselor_name = counselor.name
                    counselor.delete()
                    messages.success(request, f'Counselor "{counselor_name}" deleted successfully.')
                except Counselor.DoesNotExist:
                    messages.error(request, 'Counselor not found.')
        
        return redirect('manage_counselors')
    
    # GET request - display the management page
    from .models import Counselor
    counselors = Counselor.objects.all().order_by('name')
    active_count = counselors.filter(is_active=True).count()
    inactive_count = counselors.filter(is_active=False).count()
    
    return render(request, 'maintenance/manage_counselors.html', {
        'counselors': counselors,
        'active_count': active_count,
        'inactive_count': inactive_count,
    })

@login_required
def manage_students(request):
    if request.user.role not in ['guidance', 'maintenance', 'counselor']:
        return redirect('dashboard')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_student':
            username = request.POST.get('username', '').strip()
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            grade_level = request.POST.get('grade_level', '').strip()
            section = request.POST.get('section', '').strip()
            password = request.POST.get('password', '').strip()
            
            if username and first_name and last_name and password:
                if not CustomUser.objects.filter(username=username).exists():
                    student = CustomUser.objects.create_user(
                        username=username,
                        password=password,
                        first_name=first_name.title(),
                        last_name=last_name.title(),
                        email=email,
                        role='student',
                        grade_level=grade_level,
                        section=section,
                        is_active=True
                    )
                    messages.success(request, f'Student "{student.get_full_name()}" added successfully.')
                else:
                    messages.error(request, f'Username "{username}" already exists.')
            else:
                messages.error(request, 'Please fill in all required fields.')
                
        elif action == 'edit_student':
            student_id = request.POST.get('student_id')
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            grade_level = request.POST.get('grade_level', '').strip()
            section = request.POST.get('section', '').strip()
            is_active = request.POST.get('is_active') == 'on'
            
            if student_id:
                try:
                    student = CustomUser.objects.get(id=student_id, role='student')
                    student.first_name = first_name.title()
                    student.last_name = last_name.title()
                    student.email = email
                    student.grade_level = grade_level
                    student.section = section
                    student.is_active = is_active
                    student.save()
                    messages.success(request, f'Student "{student.get_full_name()}" updated successfully.')
                except CustomUser.DoesNotExist:
                    messages.error(request, 'Student not found.')
            else:
                messages.error(request, 'Invalid student data.')
        
        return redirect('manage_students')
    
    # GET request - display the management page
    students = CustomUser.objects.filter(role='student').order_by('last_name', 'first_name')
    
    # Add violation count to each student
    for student in students:
        student.violation_count = IncidentReport.objects.filter(reported_student=student).count()
    
    # Calculate statistics
    total_students = students.count()
    active_students = students.filter(is_active=True).count()
    inactive_students = students.filter(is_active=False).count()
    students_with_violations = students.filter(incident_reports__isnull=False).distinct().count()
    
    return render(request, 'maintenance/manage_students.html', {
        'students': students,
        'total_students': total_students,
        'active_students': active_students,
        'inactive_students': inactive_students,
        'students_with_violations': students_with_violations,
    })

@login_required
def backup_restore(request):
    if request.user.role not in ['guidance', 'maintenance', 'counselor']:
        return redirect('dashboard')
    
    backups = SystemBackup.objects.all().order_by('-created_at')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'backup':
            # Create backup (simplified)
            backup = SystemBackup.objects.create(
                backup_name=f"Manual Backup {timezone.now().strftime('%Y%m%d_%H%M%S')}",
                backup_type='manual',
                file_path=f"/backups/backup_{timezone.now().strftime('%Y%m%d_%H%M%S')}.sql",
                file_size=0,  # Would be calculated in real implementation
                created_by=request.user
            )
            messages.success(request, 'Backup created successfully')
    
    return render(request, 'maintenance/backup_restore.html', {'backups': backups})
# Analytics API Views
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
import json
import csv
import io
# PDF and Excel libraries - install with: pip install reportlab xlsxwriter
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False

@login_required
def dashboard_analytics_api(request):
    """API endpoint for dashboard analytics data"""
    user_role = request.user.role
    
    # Get date range for trends (last 12 months)
    end_date = timezone.now()
    start_date = end_date - timedelta(days=365)
    
    # Trend data - Monthly violation reports
    trend_data = []
    for i in range(12):
        month_start = start_date + timedelta(days=30*i)
        month_end = month_start + timedelta(days=30)
        month_reports = IncidentReport.objects.filter(
            created_at__range=[month_start, month_end]
        ).count()
        trend_data.append({
            'month': month_start.strftime('%b'),
            'reports': month_reports
        })
    
    # Grade level data
    grade_data = []
    for grade in range(7, 13):
        count = IncidentReport.objects.filter(grade_level=str(grade)).count()
        grade_data.append({
            'grade': f'Grade {grade}',
            'count': count
        })
    
    # Violation type distribution
    violation_types = IncidentType.objects.annotate(
        count=Count('incidentreport')
    ).values('name', 'count')
    
    violation_type_data = []
    for vtype in violation_types:
        if vtype['count'] > 0:
            violation_type_data.append({
                'name': vtype['name'],
                'value': vtype['count']
            })
    
    # Major vs Minor offenses by month
    stacked_data = []
    for i in range(12):
        month_start = start_date + timedelta(days=30*i)
        month_end = month_start + timedelta(days=30)
        
        prohibited_count = IncidentReport.objects.filter(
            created_at__range=[month_start, month_end],
            incident_type__severity='prohibited'
        ).count()
        
        school_policy_count = IncidentReport.objects.filter(
            created_at__range=[month_start, month_end],
            incident_type__severity='school_policy'
        ).count()
        
        stacked_data.append({
            'month': month_start.strftime('%b'),
            'prohibited': prohibited_count,
            'school_policy': school_policy_count
        })
    
    # Resolution data (for counselor and principal)
    resolution_data = []
    if user_role in ['counselor', 'principal']:
        for i in range(12):
            month_start = start_date + timedelta(days=30*i)
            month_end = month_start + timedelta(days=30)
            
            resolved_count = IncidentReport.objects.filter(
                created_at__range=[month_start, month_end],
                status='resolved'
            ).count()
            
            pending_count = IncidentReport.objects.filter(
                created_at__range=[month_start, month_end],
                status__in=['pending', 'under_review', 'classified', 'evaluated']
            ).count()
            
            resolution_data.append({
                'month': month_start.strftime('%b'),
                'resolved': resolved_count,
                'pending': pending_count
            })
    
    return JsonResponse({
        'trend_data': trend_data,
        'grade_data': grade_data,
        'violation_type_data': violation_type_data,
        'stacked_data': stacked_data,
        'resolution_data': resolution_data,
    })

@login_required
def export_report_api(request):
    """API endpoint for exporting reports"""
    format_type = request.GET.get('format', 'pdf')
    user_role = request.user.role
    
    # Get report data
    reports = IncidentReport.objects.all().order_by('-created_at')
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="sirms_report_{timezone.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Case ID', 'Student', 'Incident Type', 'Grade', 'Status', 'Date', 'Reporter'])
        
        for report in reports:
            writer.writerow([
                report.case_id,
                report.reported_student.get_full_name() if report.reported_student else 'N/A',
                report.incident_type.name if report.incident_type else 'N/A',
                report.grade_level,
                report.get_status_display(),
                report.created_at.strftime('%Y-%m-%d'),
                report.reporter.get_full_name() if report.reporter else 'N/A'
            ])
        
        return response
    
    elif format_type == 'excel':
        if not XLSXWRITER_AVAILABLE:
            return JsonResponse({'error': 'Excel export not available. Please install xlsxwriter: pip install xlsxwriter'}, status=400)
            
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('SIRMS Report')
        
        # Add headers
        headers = ['Case ID', 'Student', 'Incident Type', 'Grade', 'Status', 'Date', 'Reporter']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)
        
        # Add data
        for row, report in enumerate(reports, 1):
            worksheet.write(row, 0, report.case_id)
            worksheet.write(row, 1, report.reported_student.get_full_name() if report.reported_student else 'N/A')
            worksheet.write(row, 2, report.incident_type.name if report.incident_type else 'N/A')
            worksheet.write(row, 3, report.grade_level)
            worksheet.write(row, 4, report.get_status_display())
            worksheet.write(row, 5, report.created_at.strftime('%Y-%m-%d'))
            worksheet.write(row, 6, report.reporter.get_full_name() if report.reporter else 'N/A')
        
        workbook.close()
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="sirms_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
    
    elif format_type == 'pdf':
        if not REPORTLAB_AVAILABLE:
            return JsonResponse({'error': 'PDF export not available. Please install reportlab: pip install reportlab'}, status=400)
            
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="sirms_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
        
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # Title
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, "SIRMS Analytics Report")
        
        # Date
        p.setFont("Helvetica", 12)
        p.drawString(50, height - 80, f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
        
        # Summary statistics
        y_position = height - 120
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y_position, "Summary Statistics")
        
        y_position -= 30
        p.setFont("Helvetica", 12)
        total_reports = reports.count()
        pending_reports = reports.filter(status='pending').count()
        resolved_reports = reports.filter(status='resolved').count()
        
        p.drawString(50, y_position, f"Total Reports: {total_reports}")
        y_position -= 20
        p.drawString(50, y_position, f"Pending Reports: {pending_reports}")
        y_position -= 20
        p.drawString(50, y_position, f"Resolved Reports: {resolved_reports}")
        
        # Recent reports table
        y_position -= 50
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y_position, "Recent Reports")
        
        y_position -= 30
        p.setFont("Helvetica", 10)
        
        # Table headers
        headers = ['Case ID', 'Student', 'Type', 'Status', 'Date']
        x_positions = [50, 120, 250, 350, 450]
        
        for i, header in enumerate(headers):
            p.drawString(x_positions[i], y_position, header)
        
        y_position -= 20
        
        # Table data (first 20 reports)
        for report in reports[:20]:
            if y_position < 100:  # Start new page if needed
                p.showPage()
                y_position = height - 50
            
            p.drawString(x_positions[0], y_position, report.case_id)
            p.drawString(x_positions[1], y_position, 
                        (report.reported_student.get_full_name()[:15] + '...') 
                        if report.reported_student and len(report.reported_student.get_full_name()) > 15 
                        else (report.reported_student.get_full_name() if report.reported_student else 'N/A'))
            p.drawString(x_positions[2], y_position, 
                        (report.incident_type.name[:15] + '...') 
                        if report.incident_type and len(report.incident_type.name) > 15 
                        else (report.incident_type.name if report.incident_type else 'N/A'))
            p.drawString(x_positions[3], y_position, report.get_status_display())
            p.drawString(x_positions[4], y_position, report.created_at.strftime('%Y-%m-%d'))
            
            y_position -= 15
        
        p.save()
        buffer.seek(0)
        response.write(buffer.getvalue())
        buffer.close()
        
        return response
    
    return JsonResponse({'error': 'Invalid format'}, status=400)

def test_charts(request):
    """Test page to verify Chart.js is working"""
    return render(request, 'test_charts.html')

def simple_chart_test(request):
    """Simple Chart.js diagnostic test"""
    return render(request, 'simple_chart_test.html')


# Export Analytics Data
@login_required
def export_analytics(request):
    """Export analytics data in PDF, Excel, or CSV format"""
    import csv
    from django.http import HttpResponse
    from datetime import datetime, timedelta
    from django.db.models import Count, Q
    from django.db.models.functions import TruncMonth
    
    format_type = request.GET.get('format', 'csv')
    user_role = request.GET.get('role', request.user.role)
    date_range = request.GET.get('range', 'monthly')
    
    # Calculate date range
    today = datetime.now().date()
    if date_range == 'monthly':
        start_date = today - timedelta(days=30)
    elif date_range == 'quarterly':
        start_date = today - timedelta(days=90)
    else:  # yearly
        start_date = today - timedelta(days=365)
    
    # Get reports based on role
    reports = IncidentReport.objects.filter(created_at__date__gte=start_date)
    
    if user_role == 'do':
        # DO sees all reports
        pass
    elif user_role == 'counselor':
        # Counselor sees major cases
        reports = reports.filter(severity='major')
    elif user_role == 'principal':
        # Principal sees all reports
        pass
    
    # Prepare data
    data = []
    for report in reports:
        data.append({
            'Case ID': report.case_id,
            'Date': report.created_at.strftime('%Y-%m-%d'),
            'Type': report.incident_type.name if report.incident_type else 'N/A',
            'Severity': report.severity or 'N/A',
            'Status': report.status,
            'Grade': report.grade_level or 'N/A',
            'Reporter': f"{report.reporter_first_name} {report.reporter_last_name}",
        })
    
    # Export based on format
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="analytics_{user_role}_{today}.csv"'
        
        if data:
            writer = csv.DictWriter(response, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return response
    
    elif format_type == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{user_role.upper()} Analytics"
            
            # Add header
            if data:
                headers = list(data[0].keys())
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Add data
                for row_num, row_data in enumerate(data, 2):
                    for col_num, value in enumerate(row_data.values(), 1):
                        ws.cell(row=row_num, column=col_num, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="analytics_{user_role}_{today}.xlsx"'
            
            return response
        except ImportError:
            messages.error(request, 'Excel export requires openpyxl library. Please install it.')
            return redirect('dashboard')
    
    elif format_type == 'pdf':
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from io import BytesIO
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#2E8B57'),
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # Title
            title = Paragraph(f"SIRMS Analytics Report - {user_role.upper()}", title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.3*inch))
            
            # Date range info
            info_style = styles['Normal']
            info = Paragraph(f"Report Period: {start_date} to {today}", info_style)
            elements.append(info)
            elements.append(Spacer(1, 0.2*inch))
            
            # Table data
            if data:
                table_data = [list(data[0].keys())]  # Headers
                for row in data:
                    table_data.append(list(row.values()))
                
                # Create table
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E8B57')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                ]))
                
                elements.append(table)
            else:
                no_data = Paragraph("No data available for the selected period.", styles['Normal'])
                elements.append(no_data)
            
            # Build PDF
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="analytics_{user_role}_{today}.pdf"'
            
            return response
        except ImportError:
            messages.error(request, 'PDF export requires reportlab library. Please install it.')
            return redirect('dashboard')
    
    return redirect('dashboard')


# Chart testing views
@login_required
def test_charts(request):
    """Test page for chart functionality"""
    return render(request, 'test_charts.html')

@login_required
def simple_chart_test(request):
    """Simple chart test page for diagnostics"""
    return render(request, 'simple_chart_test.html')


# ESP Teacher / VPF views
@login_required
def vpf_cases(request):
    """VPF Cases management for ESP Teachers - Only show cases assigned to them"""
    if request.user.role != 'esp_teacher':
        return redirect('dashboard')
    
    from .models import VPFCase, Counselor
    
    # Find the Counselor record that matches this ESP teacher's name
    # This links the user account to the counselor in "Manage ESP Teacher/VPF"
    esp_teacher_name = request.user.get_full_name()
    matching_counselors = Counselor.objects.filter(name__icontains=esp_teacher_name)
    
    # Get only VPF cases assigned to this ESP teacher
    if matching_counselors.exists():
        vpf_cases = VPFCase.objects.filter(
            esp_teacher_assigned__in=matching_counselors
        ).select_related(
            'student', 'report', 'assigned_by', 'esp_teacher_assigned'
        ).order_by('-assigned_at')
    else:
        # No matching counselor found, show empty list
        vpf_cases = VPFCase.objects.none()
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        vpf_cases = vpf_cases.filter(status=status_filter)
    
    # Statistics
    total_cases = vpf_cases.count()
    pending_cases = vpf_cases.filter(status='pending').count()
    scheduled_cases = vpf_cases.filter(status='scheduled').count()
    completed_cases = vpf_cases.filter(status='completed').count()
    
    return render(request, 'esp/vpf_cases.html', {
        'vpf_cases': vpf_cases,
        'total_cases': total_cases,
        'pending_cases': pending_cases,
        'scheduled_cases': scheduled_cases,
        'completed_cases': completed_cases,
        'status_filter': status_filter,
    })


@login_required
def update_vpf_status(request, vpf_id):
    """Update VPF case status by ESP Teacher"""
    if request.user.role != 'esp_teacher':
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    from .models import VPFCase
    
    vpf_case = get_object_or_404(VPFCase, id=vpf_id)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        notes = request.POST.get('notes', '')
        
        if not new_status or not notes:
            return JsonResponse({'success': False, 'message': 'Status and notes are required'}, status=400)
        
        # Update status
        old_status = vpf_case.status
        vpf_case.status = new_status
        
        # Append notes with timestamp
        timestamp = timezone.now().strftime('%Y-%m-%d %H:%M')
        status_update = f"\n\n[{timestamp}] Status updated from '{old_status}' to '{new_status}' by {request.user.get_full_name()}:\n{notes}"
        
        if vpf_case.notes:
            vpf_case.notes += status_update
        else:
            vpf_case.notes = status_update.strip()
        
        vpf_case.save()
        
        # Notify guidance counselor who assigned the case
        if vpf_case.assigned_by:
            Notification.objects.create(
                user=vpf_case.assigned_by,
                title=f'VPF Status Updated - {vpf_case.report.case_id}',
                message=f'ESP Teacher {request.user.get_full_name()} updated VPF case status to "{new_status}" for {vpf_case.student.get_full_name()}.\n\nNotes: {notes}',
                report=vpf_case.report
            )
        
        return JsonResponse({'success': True, 'message': 'Status updated successfully'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)

@login_required
def vpf_schedule(request):
    """VPF Schedule management for ESP Teachers"""
    if request.user.role != 'esp_teacher':
        return redirect('dashboard')
    
    from .models import VPFCase, VPFSchedule, Counselor
    
    if request.method == 'POST':
        from datetime import datetime, timedelta
        
        vpf_case_id = request.POST.get('vpf_case_id')
        scheduled_date_str = request.POST.get('scheduled_date')
        location = request.POST.get('location', '')
        notes = request.POST.get('notes', '')
        
        vpf_case = get_object_or_404(VPFCase, id=vpf_case_id)
        
        # Parse the datetime string
        try:
            scheduled_date = datetime.strptime(scheduled_date_str, '%Y-%m-%dT%H:%M')
        except ValueError:
            # Try alternative format
            scheduled_date = datetime.strptime(scheduled_date_str, '%Y-%m-%d %H:%M:%S')
        
        # Check for duplicate schedule (same VPF case already scheduled)
        existing_schedule = VPFSchedule.objects.filter(
            vpf_case=vpf_case,
            status__in=['scheduled', 'ongoing']
        ).first()
        
        if existing_schedule:
            messages.error(request, f'This VPF case is already scheduled for {existing_schedule.scheduled_date.strftime("%B %d, %Y at %I:%M %p")}. Please cancel or complete the existing schedule first.')
            return redirect('vpf_schedule')
        
        # Check for time conflict (same ESP teacher, overlapping time)
        time_buffer = timedelta(hours=1)  # 1 hour buffer
        conflicting_schedules = VPFSchedule.objects.filter(
            esp_teacher=request.user,
            status__in=['scheduled', 'ongoing'],
            scheduled_date__gte=scheduled_date - time_buffer,
            scheduled_date__lte=scheduled_date + time_buffer
        )
        
        if conflicting_schedules.exists():
            conflict = conflicting_schedules.first()
            messages.error(request, f'Time conflict! You already have a session scheduled at {conflict.scheduled_date.strftime("%I:%M %p")} with {conflict.vpf_case.student.get_full_name()}. Please choose a different time.')
            return redirect('vpf_schedule')
        
        # Create schedule (no counselor assignment needed)
        schedule = VPFSchedule.objects.create(
            vpf_case=vpf_case,
            esp_teacher=request.user,
            counselor_assigned=None,
            scheduled_date=scheduled_date,
            location=location,
            notes=notes,
            status='scheduled'
        )
        
        # Update VPF case status
        vpf_case.status = 'scheduled'
        vpf_case.save()
        
        # Notify student
        Notification.objects.create(
            user=vpf_case.student,
            title='VPF Session Scheduled',
            message=f'You have been scheduled for a Values Reflective Formation session on {scheduled_date.strftime("%B %d, %Y at %I:%M %p")}. Location: {location}. Please attend on time.',
            report=vpf_case.report
        )
        
        # Notify the adviser/teacher based on student's curriculum, grade, and section
        notify_adviser_of_counseling(
            report=vpf_case.report,
            scheduled_date=scheduled_date,
            location=location,
            counselor_name=request.user.get_full_name(),
            schedule_type='VPF'
        )
        
        messages.success(request, f'VPF session scheduled for {vpf_case.student.get_full_name()} on {scheduled_date.strftime("%B %d, %Y at %I:%M %p")}')
        return redirect('vpf_schedule')
    
    # Find the Counselor record that matches this ESP teacher's name
    esp_teacher_name = request.user.get_full_name()
    matching_counselors = Counselor.objects.filter(name__icontains=esp_teacher_name)
    
    # Get pending VPF cases assigned to this ESP teacher (not yet scheduled)
    if matching_counselors.exists():
        pending_vpf_cases = VPFCase.objects.filter(
            esp_teacher_assigned__in=matching_counselors,
            status='pending'
        ).select_related('student', 'report', 'esp_teacher_assigned')
    else:
        pending_vpf_cases = VPFCase.objects.none()
    
    # Get schedules for cases assigned to this ESP teacher
    if matching_counselors.exists():
        schedules = VPFSchedule.objects.filter(
            vpf_case__esp_teacher_assigned__in=matching_counselors
        ).select_related(
            'vpf_case__student', 'counselor_assigned', 'esp_teacher'
        ).order_by('-scheduled_date')
    else:
        schedules = VPFSchedule.objects.none()
    
    return render(request, 'esp/vpf_schedule.html', {
        'pending_vpf_cases': pending_vpf_cases,
        'schedules': schedules,
    })


@login_required
def for_vpf(request):
    """VPF Cases assigned by the counselor"""
    if request.user.role != 'counselor':
        return redirect('dashboard')
    
    from .models import VPFCase, Counselor
    
    # Get VPF cases assigned by this counselor
    vpf_cases = VPFCase.objects.filter(
        assigned_by=request.user
    ).select_related(
        'student', 'report', 'esp_teacher_assigned'
    ).order_by('-assigned_at')
    
    # Get VPF teachers from Counselor model (Manage Counselors)
    vpf_teachers = Counselor.objects.filter(is_active=True).order_by('name')
    
    # Statistics
    total_cases = vpf_cases.count()
    pending_cases = vpf_cases.filter(status='pending').count()
    scheduled_cases = vpf_cases.filter(status='scheduled').count()
    completed_cases = vpf_cases.filter(status='completed').count()
    
    return render(request, 'counselor/for_vpf.html', {
        'vpf_cases': vpf_cases,
        'vpf_teachers': vpf_teachers,
        'total_cases': total_cases,
        'pending_cases': pending_cases,
        'scheduled_cases': scheduled_cases,
        'completed_cases': completed_cases,
    })


@login_required
def assign_vpf_teacher(request):
    """Assign ESP Teacher to VPF Case"""
    if request.user.role != 'counselor':
        return redirect('dashboard')
    
    if request.method == 'POST':
        from .models import VPFCase, Counselor
        
        vpf_case_id = request.POST.get('vpf_case_id')
        teacher_id = request.POST.get('teacher_id')
        assignment_notes = request.POST.get('assignment_notes', '')
        
        try:
            vpf_case = get_object_or_404(VPFCase, id=vpf_case_id, assigned_by=request.user)
            teacher = get_object_or_404(Counselor, id=teacher_id)
            
            # Assign teacher to VPF case
            vpf_case.esp_teacher_assigned = teacher
            if assignment_notes:
                vpf_case.notes = f"{vpf_case.notes}\n\nAssignment Notes: {assignment_notes}" if vpf_case.notes else f"Assignment Notes: {assignment_notes}"
            vpf_case.save()
            
            # Notify ESP teachers (users with esp_teacher role)
            esp_teachers = CustomUser.objects.filter(role='esp_teacher', is_active=True)
            for esp in esp_teachers:
                Notification.objects.create(
                    user=esp,
                    title='VPF Case Assigned to Teacher',
                    message=f'Case {vpf_case.report.case_id} for student {vpf_case.student.get_full_name()} has been assigned to {teacher.name}. Please schedule the VPF session.',
                    report=vpf_case.report
                )
            
            # Notify the student
            Notification.objects.create(
                user=vpf_case.student,
                title='VPF Teacher Assigned',
                message=f'Your VPF case ({vpf_case.report.case_id}) has been assigned to {teacher.name}. You will be notified when the session is scheduled.',
                report=vpf_case.report
            )
            
            messages.success(request, f'ESP Teacher {teacher.name} assigned successfully to case {vpf_case.report.case_id}. Notifications sent.')
            
        except Exception as e:
            messages.error(request, f'Error assigning teacher: {str(e)}')
    
    return redirect('for_vpf')


@login_required
def manage_esp_teachers(request):
    """Manage ESP Teachers/VPF Counselors"""
    if request.user.role != 'counselor':
        return redirect('dashboard')
    
    from .models import Counselor
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            name = request.POST.get('name')
            email = request.POST.get('email', '')
            phone = request.POST.get('phone', '')
            specialization = request.POST.get('specialization', '')
            is_active = request.POST.get('is_active') == 'on'
            
            Counselor.objects.create(
                name=name,
                email=email,
                phone=phone,
                specialization=specialization,
                is_active=is_active
            )
            messages.success(request, f'ESP Teacher "{name}" added successfully.')
            
        elif action == 'edit':
            teacher_id = request.POST.get('teacher_id')
            teacher = get_object_or_404(Counselor, id=teacher_id)
            
            teacher.name = request.POST.get('name')
            teacher.email = request.POST.get('email', '')
            teacher.phone = request.POST.get('phone', '')
            teacher.specialization = request.POST.get('specialization', '')
            teacher.is_active = request.POST.get('is_active') == 'on'
            teacher.save()
            
            messages.success(request, f'ESP Teacher "{teacher.name}" updated successfully.')
            
        elif action == 'delete':
            teacher_id = request.POST.get('teacher_id')
            teacher = get_object_or_404(Counselor, id=teacher_id)
            teacher_name = teacher.name
            teacher.delete()
            
            messages.success(request, f'ESP Teacher "{teacher_name}" deleted successfully.')
        
        return redirect('manage_esp_teachers')
    
    # GET request - display list
    esp_teachers = Counselor.objects.all().order_by('name')
    
    return render(request, 'counselor/manage_esp_teachers.html', {
        'esp_teachers': esp_teachers
    })


@login_required
def get_dashboard_analytics(request):
    """API endpoint for dynamic dashboard analytics based on time filter"""
    from django.http import JsonResponse
    from datetime import datetime, timedelta
    from django.utils import timezone
    from django.db.models import Count
    import json
    
    time_filter = request.GET.get('filter', 'monthly')
    user = request.user
    
    # Calculate date range based on filter
    end_date = timezone.now()
    if time_filter == 'monthly':
        start_date = end_date - timedelta(days=365)  # Last 12 months
        periods = 12
        period_days = 30
        date_format = '%b'
    elif time_filter == 'quarterly':
        start_date = end_date - timedelta(days=365*2)  # Last 8 quarters
        periods = 8
        period_days = 90
        date_format = 'Q%m'
    else:  # yearly
        start_date = end_date - timedelta(days=365*5)  # Last 5 years
        periods = 5
        period_days = 365
        date_format = '%Y'
    
    # Generate trend data
    trend_data = []
    for i in range(periods):
        period_start = start_date + timedelta(days=period_days*i)
        period_end = period_start + timedelta(days=period_days)
        
        period_reports = IncidentReport.objects.filter(
            created_at__range=[period_start, period_end]
        ).count()
        
        # Format label based on filter type
        if time_filter == 'quarterly':
            quarter = (period_start.month - 1) // 3 + 1
            label = f'Q{quarter} {period_start.year}'
        else:
            label = period_start.strftime(date_format)
        
        trend_data.append({
            'month': label,
            'reports': period_reports
        })
    
    # Generate grade data (always the same regardless of time filter)
    grade_data = []
    for grade in range(7, 13):
        count = IncidentReport.objects.filter(
            grade_level=str(grade),
            created_at__gte=start_date
        ).count()
        grade_data.append({
            'grade': f'Grade {grade}',
            'count': count
        })
    
    # Generate stacked data (Prohibited Acts vs Other School Policies)
    stacked_data = []
    for i in range(periods):
        period_start = start_date + timedelta(days=period_days*i)
        period_end = period_start + timedelta(days=period_days)
        
        # Prohibited Acts (severity='prohibited')
        prohibited_count = IncidentReport.objects.filter(
            created_at__range=[period_start, period_end],
            incident_type__severity='prohibited'
        ).count()
        
        # Other School Policies (severity='school_policy')
        school_policy_count = IncidentReport.objects.filter(
            created_at__range=[period_start, period_end],
            incident_type__severity='school_policy'
        ).count()
        
        # Format label
        if time_filter == 'quarterly':
            quarter = (period_start.month - 1) // 3 + 1
            label = f'Q{quarter} {period_start.year}'
        else:
            label = period_start.strftime(date_format)
        
        stacked_data.append({
            'month': label,
            'prohibited': prohibited_count,
            'school_policy': school_policy_count
        })
    
    # Calculate metrics based on role
    metrics = {}
    
    if user.role == 'do':
        reports = IncidentReport.objects.filter(created_at__gte=start_date)
        minor_cases = reports.filter(classification__severity='minor').count()
        major_cases = reports.filter(classification__severity='major').count()
        
        metrics = {
            'total_reports': reports.count(),
            'pending': reports.filter(status='pending').count(),
            'minor_cases_count': minor_cases,
            'major_cases_count': major_cases,
        }
    
    elif user.role == 'counselor':
        reports = IncidentReport.objects.filter(created_at__gte=start_date)
        
        # Resolution data for counselor
        resolution_data = []
        for i in range(periods):
            period_start = start_date + timedelta(days=period_days*i)
            period_end = period_start + timedelta(days=period_days)
            
            resolved_count = IncidentReport.objects.filter(
                created_at__range=[period_start, period_end],
                status='resolved'
            ).count()
            
            pending_count = IncidentReport.objects.filter(
                created_at__range=[period_start, period_end],
                status__in=['pending', 'under_review', 'classified', 'evaluated']
            ).count()
            
            # Format label
            if time_filter == 'quarterly':
                quarter = (period_start.month - 1) // 3 + 1
                label = f'Q{quarter} {period_start.year}'
            else:
                label = period_start.strftime(date_format)
            
            resolution_data.append({
                'month': label,
                'resolved': resolved_count,
                'pending': pending_count
            })
        
        metrics = {
            'total_prohibited_acts': reports.filter(incident_type__severity='prohibited').count(),
            'total_osp': reports.filter(incident_type__severity='school_policy').count(),
            'scheduled_sessions': CounselingSchedule.objects.filter(
                counselor=user,
                status='scheduled',
                created_at__gte=start_date
            ).count(),
            'completed_sessions': CounselingSchedule.objects.filter(
                counselor=user,
                status='completed',
                created_at__gte=start_date
            ).count(),
            'resolution_data': resolution_data,
        }
    
    elif user.role == 'esp_teacher':
        from .models import VPFCase, Counselor
        
        # Find matching counselor record
        esp_teacher_name = user.get_full_name()
        matching_counselors = Counselor.objects.filter(name__icontains=esp_teacher_name)
        
        if matching_counselors.exists():
            vpf_cases = VPFCase.objects.filter(
                esp_teacher_assigned__in=matching_counselors,
                assigned_at__gte=start_date
            )
        else:
            vpf_cases = VPFCase.objects.filter(assigned_at__gte=start_date)
        
        metrics = {
            'total_vpf_referrals': vpf_cases.count(),
            'completed_vpf': vpf_cases.filter(status='completed').count(),
            'pending_vpf': vpf_cases.filter(status='pending').count(),
            'ongoing_vpf': vpf_cases.filter(status='ongoing').count(),
        }
    
    response_data = {
        'trend_data': trend_data,
        'grade_data': grade_data,
        'stacked_data': stacked_data,
        'metrics': metrics,
    }
    
    return JsonResponse(response_data)
